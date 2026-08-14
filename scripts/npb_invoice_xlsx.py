#!/usr/bin/env python3
"""실비 청구서 xlsx 생성.

26년 6월분부터 운임/광고 실비는 판매정산서에 합치지 않고 따로 발행한다.
공헌이익을 볼 때 계정별로 갈라 보기 위해서다. 그래서 정산서와 같은 파일에
붙이지 않고 청구서 한 장으로 독립시킨다.

운임/물류는 시트가 여러 장이다.
    시트1  운임비용내역서 — 비용집계, 3PL 단가표, 메모/특이사항
    시트2~ 올린 출고내역 파일 원본 (유형별 한 장씩)
단가표와 문구는 브랜드 설정에서 내려온 값을 그대로 쓴다 — 여기에 값을 적어
두면 설정을 고쳐도 청구서가 안 바뀐다.

    python3 npb_invoice_xlsx.py --input <invoice.json> --output <out.xlsx>
"""

import argparse
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="D0D0D0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
# 색은 알파까지 적는다 — 접두를 빼면 00 으로 저장돼 양식 파일과 값이 달라진다.
HEAD_FILL = PatternFill("solid", fgColor="FFF2F2F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFFFF6E5")
BILL_FILL = PatternFill("solid", fgColor="FFFDE9D9")
MONEY = "#,##0"

WRAP = Alignment(wrap_text=True, vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
VCENTER = Alignment(vertical="center")

# 양식의 본문 글자 크기.
BASE_FONT = 10


def put(ws, row, col, value, *, bold=False, fill=None, box=True,
        number=False, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(size=BASE_FONT, bold=bold)
    if fill:
        cell.fill = fill
    if box:
        cell.border = BOX
    if number:
        cell.number_format = MONEY
        cell.alignment = RIGHT
    if align:
        cell.alignment = align
    return cell


def write_rows(ws, rows, start=1):
    """업로드 원본을 그대로 적는다.

    파서는 행을 {열이름: 값} 으로 준다. 열 순서는 첫 행 기준으로 잡되, 뒤쪽
    행에만 있는 열도 빠뜨리지 않도록 나오는 순서대로 이어 붙인다.
    """
    if not rows:
        return
    if isinstance(rows[0], dict):
        header = []
        seen = set()
        for row in rows:
            for key in row.keys():
                if key not in seen:
                    seen.add(key)
                    header.append(key)
        for c, key in enumerate(header, start=1):
            cell = ws.cell(row=start, column=c, value=key)
            cell.font = Font(bold=True)
            cell.fill = HEAD_FILL
        for r, row in enumerate(rows, start=start + 1):
            for c, key in enumerate(header, start=1):
                ws.cell(row=r, column=c, value=row.get(key))
        return
    for r, row in enumerate(rows, start=start):
        for c, value in enumerate(row or [], start=1):
            ws.cell(row=r, column=c, value=value)


def band(ws, row, first, last, fill=None):
    """한 줄 전체에 테두리(있으면 배경)를 입힌다."""
    for col in range(first, last + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = BOX
        if fill:
            cell.fill = fill


def build_logistics(wb, invoice, sheet):
    """확정된 양식(픽키도기클럽 2026-06 사용내역서)과 셀 단위로 같게 그린다.

    행 위치가 달마다 흔들리지 않도록 블록 순서를 고정한다. 유일하게 늘어나는
    곳은 단가표이고, 그 아래 블록만 밀린다.
    """
    ws = wb.active
    ws.title = "운임비용내역서"
    for col, width in [("A", 2.0), ("B", 20.5), ("C", 14.0), ("D", 14.0),
                       ("E", 20.0), ("F", 24.7), ("G", 18.7)]:
        ws.column_dimensions[col].width = width

    # --- 제목 ---
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 33
    title = ws.cell(row=2, column=2, value=sheet.get("title") or "운송/물류 내역 확인 & 청구서")
    title.font = Font(size=16, bold=True)
    title.alignment = CENTER
    ws.merge_cells("B2:G2")
    issuer = ws.cell(row=3, column=7, value=sheet.get("issuerName") or "")
    issuer.alignment = RIGHT

    # --- 머리 정보 ---
    row = 4
    for label, value in [
        ("집계기간", invoice.get("periodMonth") or ""),
        ("집계기준", sheet.get("basisLabel") or ""),
        ("브랜드", invoice.get("brandName") or ""),
        ("발행일", str(invoice.get("issuedAt") or "")[:10]),
        ("입금요청", sheet.get("paymentTerms") or invoice.get("dueDate") or ""),
    ]:
        put(ws, row, 2, label, bold=True, fill=HEAD_FILL, align=VCENTER)
        put(ws, row, 3, value, align=VCENTER)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        band(ws, row, 3, 7)
        row += 1

    # --- 비용집계 ---
    row += 1
    put(ws, row, 2, "[실비 : 운임/물류]", bold=True, box=False)
    row += 1
    head_row = row
    put(ws, row, 2, "정산 월", bold=True, fill=HEAD_FILL, align=VCENTER)
    put(ws, row, 3, "물류 실비 산정", bold=True, fill=HEAD_FILL, align=CENTER)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    band(ws, row, 3, 6, HEAD_FILL)
    put(ws, row, 7, "비고", bold=True, fill=HEAD_FILL, align=CENTER)
    ws.merge_cells(start_row=row, start_column=7, end_row=row + 1, end_column=7)
    row += 1
    put(ws, row, 2, "", bold=True, fill=HEAD_FILL)
    for col, label in [(3, "수량(출고건수)"), (4, "운임비"),
                       (5, "피킹/패킹/부자재"), (6, "총계")]:
        put(ws, row, col, label, bold=True, fill=HEAD_FILL, align=CENTER)
    band(ws, row, 7, 7, HEAD_FILL)
    row += 1

    for item in sheet.get("rows") or []:
        put(ws, row, 2, item.get("label") or "", align=VCENTER)
        put(ws, row, 3, item.get("count") or 0, number=True)
        put(ws, row, 4, item.get("freight") or 0, number=True)
        put(ws, row, 5, item.get("handling") or 0, number=True)
        put(ws, row, 6, item.get("amount") or 0, number=True)
        put(ws, row, 7, item.get("note") or "", align=VCENTER)
        row += 1

    # 용달·퀵은 건마다 금액이 달라 합계에 넣지 않는다. 없는 달도 줄은 남긴다 —
    # 양식이 달마다 같은 모양이어야 대조하기 쉽다.
    for item in sheet.get("separateRows") or []:
        put(ws, row, 2, item.get("label") or "", align=VCENTER)
        put(ws, row, 3, item.get("count") or None, number=True)
        put(ws, row, 4, item.get("amount") or None, number=True)
        put(ws, row, 5, None)
        put(ws, row, 6, item.get("amount") or None, number=True)
        put(ws, row, 7, item.get("note") or "", align=VCENTER)
        row += 1

    put(ws, row, 2, "합계", bold=True, fill=TOTAL_FILL, align=VCENTER)
    put(ws, row, 3, sheet.get("countTotal") or 0, number=True, fill=TOTAL_FILL)
    put(ws, row, 4, None, fill=TOTAL_FILL)
    put(ws, row, 5, None, fill=TOTAL_FILL)
    cell = put(ws, row, 6, sheet.get("subtotal") or 0, number=True, fill=TOTAL_FILL)
    cell.font = Font(bold=True)
    put(ws, row, 7, "vat포함", align=RIGHT)
    row += 1

    # 용달/퀵이 실제로 발생한 달만 청구액 줄을 세운다. 합계가 3PL·본사만
    # 세므로, 그 줄이 없으면 청구서가 실제 청구액보다 적게 적힌다.
    if sheet.get("separateTotal"):
        put(ws, row, 2, "청구액", bold=True, fill=BILL_FILL, align=VCENTER)
        put(ws, row, 3, None, fill=BILL_FILL)
        put(ws, row, 4, None, fill=BILL_FILL)
        put(ws, row, 5, None, fill=BILL_FILL)
        cell = put(ws, row, 6, sheet.get("billed") or 0, number=True, fill=BILL_FILL)
        cell.font = Font(bold=True)
        put(ws, row, 7, "합계 + 용달/퀵", align=RIGHT)
        row += 1

    # --- 표 아래 주석 ---
    row += 1
    notes = sheet.get("footnotes") or []
    for i, text in enumerate(notes):
        cell = ws.cell(row=row, column=2, value=f"*{text}")
        cell.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.row_dimensions[row].height = 41 if "\n" in str(text) else 28
        # 주석 묶음을 한 상자로 두른다.
        for col in range(2, 8):
            c = ws.cell(row=row, column=col)
            c.border = Border(left=THIN, right=THIN,
                              bottom=THIN if i == len(notes) - 1 else None)
        row += 1

    # --- 단가표 (vat 별도) ---
    row += 1
    put(ws, row, 7, "vat별도", box=False, align=RIGHT)
    row += 1
    for col, label in [(2, "구분"), (3, "항목"), (4, "견적(원)"), (5, "단위"), (6, "비고")]:
        put(ws, row, col, label, bold=True, fill=HEAD_FILL, align=CENTER)
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
    band(ws, row, 6, 7, HEAD_FILL)
    row += 1

    # 구분은 같은 묶음끼리 세로로 합친다.
    groups = []
    for entry in sheet.get("priceTable") or []:
        group = entry.get("group") or ""
        if groups and groups[-1][0] == group:
            groups[-1][1].append(entry)
        else:
            groups.append((group, [entry]))
    for group, entries in groups:
        first = row
        for entry in entries:
            put(ws, row, 2, group if row == first else None, bold=True, align=CENTER)
            put(ws, row, 3, entry.get("item") or None, align=VCENTER)
            label = entry.get("unitPriceLabel")
            if label:
                put(ws, row, 4, label, align=RIGHT)
            elif entry.get("unitPrice") is None:
                put(ws, row, 4, "-", align=RIGHT)
            else:
                put(ws, row, 4, entry.get("unitPrice"), number=True)
            put(ws, row, 5, entry.get("unit") or None, align=CENTER)
            cell = put(ws, row, 6, entry.get("note") or None)
            cell.alignment = WRAP
            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
            band(ws, row, 6, 7)
            if len(str(entry.get("note") or "")) > 60:
                ws.row_dimensions[row].height = 47
            row += 1
        if len(entries) > 1:
            ws.merge_cells(start_row=first, start_column=2, end_row=row - 1, end_column=2)

    # --- 메모/특이사항 ---
    row += 1
    put(ws, row, 2, "[메모/특이사항]", bold=True, box=False)
    row += 1
    for text in sheet.get("memos") or []:
        cell = ws.cell(row=row, column=2, value=text)
        cell.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 1

    # --- 올린 출고내역 원본을 유형별로 한 장씩 ---
    for source in sheet.get("sources") or []:
        title = str(source.get("label") or source.get("shipType") or "출고내역")[:28]
        sh = wb.create_sheet(title=title)
        if not source.get("rows"):
            continue
        sh.cell(row=1, column=1, value=source.get("fileName") or "").font = Font(bold=True)
        sh.cell(row=2, column=1,
                value=f"{source.get('basisLabel') or ''} 기준 {source.get('autoCount') or 0}건")
        write_rows(sh, source.get("rows") or [], start=4)


def attach_source_tabs(wb, wanted):
    """구글시트 탭을 청구서 뒤에 붙일 모양으로 정리한다.

    양식은 탭 이름 앞에 DB_ 를 붙이고 필요한 탭만 순서대로 둔다. 이름을 바꾸면
    탭끼리 참조하는 수식('월별 네이버…'!C2)이 가리킬 곳을 잃으므로, 이름을
    바꾼 뒤 수식의 참조도 같이 고쳐 준다.
    """
    keep = [name for name in (wanted or []) if name in wb.sheetnames]
    if keep:
        for name in wb.sheetnames:
            if name not in keep:
                del wb[name]
    else:
        keep = list(wb.sheetnames)

    renames = {}
    for name in keep:
        new = name if name.startswith("DB_") else f"DB_{name}"
        renames[name] = new[:31]
    for old, new in renames.items():
        wb[old].title = new
    # 시트 순서를 양식과 같게 둔다.
    wb._sheets = [wb[renames[name]] for name in keep]

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                for old, new in renames.items():
                    if old == new:
                        continue
                    value = value.replace(f"'{old}'!", f"'{new}'!")
                    value = value.replace(f"{old}!", f"'{new}'!")
                cell.value = value


def build_simple(ws, invoice):
    """광고비 등, 항목 목록만 있는 청구서. 확정된 양식과 셀 단위로 같게 그린다."""
    for col, width in [("A", 2.0), ("B", 34.0), ("C", 10.0),
                       ("D", 12.0), ("E", 14.0), ("F", 40.0)]:
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 34
    title = ws.cell(row=2, column=2, value=invoice.get("title") or "실비 청구서")
    title.font = Font(size=16, bold=True)
    title.alignment = CENTER
    ws.merge_cells("B2:F2")
    issuer = ws.cell(row=3, column=6, value=invoice.get("issuerName") or "")
    issuer.font = Font(size=BASE_FONT)
    issuer.alignment = RIGHT

    # 사업자명과 브랜드명이 다르면 둘 다 적는다 — 계산서는 사업자로 나가고
    # 내역은 브랜드로 관리한다.
    business = invoice.get("businessName") or ""
    brand = invoice.get("brandName") or ""
    if business and brand and business != brand:
        payer = f"{business}({brand})"
    else:
        payer = business or brand

    row = 4
    for label, value in [
        ("공급받는 자", payer),
        ("집계 기간", invoice.get("periodMonth") or ""),
        ("계정 분류", invoice.get("account") or ""),
        ("발행일", str(invoice.get("issuedAt") or "")[:10]),
        ("입금요청", invoice.get("paymentTerms") or invoice.get("dueDate") or ""),
    ]:
        put(ws, row, 2, label, bold=True, fill=HEAD_FILL, align=VCENTER)
        put(ws, row, 3, value, align=VCENTER)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        band(ws, row, 3, 6)
        row += 1
    row += 1

    for i, label in enumerate(["항목", "건수", "단가", "금액", "비고"], start=2):
        put(ws, row, i, label, bold=True, fill=HEAD_FILL, align=CENTER)
    row += 1

    for item in invoice.get("items") or []:
        put(ws, row, 2, item.get("label") or "", align=VCENTER)
        put(ws, row, 3, item.get("count") or None, number=True)
        put(ws, row, 4, item.get("unit") or None, number=True)
        put(ws, row, 5, item.get("amount") or 0, number=True)
        put(ws, row, 6, item.get("note") or None, align=VCENTER)
        row += 1

    put(ws, row, 2, "합계", bold=True, fill=TOTAL_FILL, align=VCENTER)
    for col in (3, 4, 6):
        put(ws, row, col, None, fill=TOTAL_FILL)
    cell = put(ws, row, 5, invoice.get("total") or 0, number=True, fill=TOTAL_FILL)
    cell.font = Font(size=BASE_FONT, bold=True)
    row += 2

    put(ws, row, 2, "[안내]", bold=True, box=False)
    ws.row_dimensions[row].height = 16
    row += 1
    notes = [
        "본 청구서는 판매정산서와 별도로 발행되는 실비 청구분입니다.",
        "정산서에는 포함·공제되지 않으며, 계정별 분류를 위해 분리 발행합니다.",
        "금액은 VAT 포함 기준입니다.",
    ]
    if invoice.get("note"):
        notes.append(str(invoice["note"]))
    for i, text in enumerate(notes):
        cell = ws.cell(row=row, column=2, value=text)
        cell.font = Font(size=BASE_FONT)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        for col in range(2, 7):
            ws.cell(row=row, column=col).border = Border(
                left=THIN, right=THIN,
                bottom=THIN if i == len(notes) - 1 else None)
        ws.row_dimensions[row].height = 16
        row += 1


def build(invoice, path, source=None):
    sheet = invoice.get("logisticsSheet")
    if sheet:
        wb = Workbook()
        build_logistics(wb, invoice, sheet)
        wb.save(path)
        return
    # 광고비는 구글시트를 통째로 받아 그 위에 청구서 장을 얹는다. 원본 탭의
    # 수식·서식이 그대로 남는다.
    if source:
        wb = load_workbook(source)
        attach_source_tabs(wb, invoice.get("sheetTabs") or [])
        ws = wb.create_sheet(title="청구서", index=0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "청구서"
    build_simple(ws, invoice)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--source", help="뒤에 붙일 원본 워크북(광고비 구글시트)")
    args = ap.parse_args()
    with open(args.input, encoding="utf-8") as fh:
        invoice = json.load(fh)
    build(invoice, args.output, args.source)


if __name__ == "__main__":
    main()
