#!/usr/bin/env python3

import argparse
import json
from copy import copy
from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook, load_workbook


UPLOAD_HEADERS = [
    "entryId",
    "처리유형",
    "브랜드명",
    "품목코드",
    "품목명",
    "옵션",
    "수량",
    "공급가",
    "원판매가",
    "할인가",
    "현재판매가",
    "적용시작일",
    "적용종료일",
    "바코드",
    "사용상태",
    "메모",
]

ACTION_MAP = {
    "": "",
    "신규": "create",
    "등록": "create",
    "create": "create",
    "수정": "update",
    "업데이트": "update",
    "update": "update",
    "개정": "revise",
    "개정추가": "revise",
    "revise": "revise",
    "삭제": "delete",
    "delete": "delete",
}


def normalize_action(value):
    text = str(value or "").strip()
    return ACTION_MAP.get(text, text.lower())


def normalize_bool(value):
    text = str(value or "").strip().lower()
    if text in {"", "y", "사용", "true", "1", "yes", "o"}:
        return True
    if text in {"n", "중지", "false", "0", "no", "x"}:
        return False
    return True


def normalize_number(value):
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    return float(text)


def normalize_date(value):
    if value in ("", None):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def export_workbook(input_path, output_path):
    payload = json.loads(Path(input_path).read_text(encoding="utf-8"))
    brand_name = payload.get("brandName", "")
    rows = payload.get("rows", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "단가업로드"
    ws.append(UPLOAD_HEADERS)

    for cell in ws[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font

    for row in rows:
        ws.append([
            row.get("entryId", ""),
            row.get("action", "수정"),
            row.get("brandName", brand_name),
            row.get("itemCode", ""),
            row.get("itemName", ""),
            row.get("spec", ""),
            row.get("unit", ""),
            row.get("supplyPrice", 0),
            row.get("originalPrice", row.get("consumerPrice", 0)),
            row.get("discountPrice", 0),
            row.get("salePrice", 0),
            row.get("effectiveFrom", ""),
            row.get("effectiveTo", ""),
            row.get("barcode", ""),
            "Y" if row.get("isActive", True) else "N",
            row.get("note", ""),
        ])

    widths = {
        "A": 22,
        "B": 12,
        "C": 28,
        "D": 18,
        "E": 32,
        "F": 18,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 14,
        "M": 14,
        "N": 18,
        "O": 10,
        "P": 28,
    }
    for key, width in widths.items():
        ws.column_dimensions[key].width = width

    guide = wb.create_sheet("작성안내")
    guide["A1"] = "작성 규칙"
    guide["A2"] = "1. 브랜드 선택 후 다운로드한 파일을 그대로 수정해서 다시 업로드합니다."
    guide["A3"] = "2. 기존 행은 entryId가 채워져 있으므로 그대로 두고 값만 수정하면 일괄 수정됩니다."
    guide["A4"] = "3. 신규 행은 entryId를 비우고 처리유형을 '신규'로 두면 됩니다."
    guide["A5"] = "4. 기존 품목의 새 단가 이력을 추가하려면 처리유형을 '개정추가'로 바꾸고 적용시작일을 새로 넣습니다."
    guide["A6"] = "5. 삭제하려면 처리유형을 '삭제'로 바꾸고 entryId는 유지합니다."
    guide["A10"] = "8. 적용종료일을 비우면 종료일 없이 상시 적용됩니다. 일시 할인/판촉 단가에 종료일을 지정하세요."
    guide["A7"] = "6. 원판매가/할인가/현재판매가를 나눠 입력할 수 있습니다. 요청 계산은 현재판매가를 기준으로 합니다."
    guide["A8"] = "7. 사용상태는 Y/N으로 입력합니다 (이전 양식의 사용/중지, true/false도 허용)."
    guide["A9"] = f"대상 브랜드: {brand_name}"
    guide.column_dimensions["A"].width = 110

    wb.save(output_path)


def import_workbook(input_path):
    wb = load_workbook(input_path, data_only=True)
    ws = wb["단가업로드"] if "단가업로드" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    headers = [str(value or "").strip() for value in header_cells]
    header_index = {name: idx for idx, name in enumerate(headers)}

    rows = []
    for excel_row in range(2, ws.max_row + 1):
        values = [ws.cell(excel_row, idx + 1).value for idx in range(len(headers))]
        raw = {name: values[idx] if idx < len(values) else "" for name, idx in header_index.items()}
        if not any(str(value or "").strip() for value in raw.values()):
            continue
        rows.append({
            "rowNumber": excel_row,
            "entryId": str(raw.get("entryId") or "").strip(),
            "action": normalize_action(raw.get("처리유형")),
            "brandName": str(raw.get("브랜드명") or "").strip(),
            "itemCode": str(raw.get("품목코드") or "").strip(),
            "itemName": str(raw.get("품목명") or "").strip(),
            "spec": str(raw.get("옵션") or raw.get("규격") or "").strip(),
            "unit": str(raw.get("수량") or raw.get("단위") or "").strip(),
            "supplyPrice": normalize_number(raw.get("공급가")),
            "originalPrice": normalize_number(raw.get("원판매가") or raw.get("소비자가")),
            "discountPrice": normalize_number(raw.get("할인가")),
            "salePrice": normalize_number(raw.get("현재판매가") or raw.get("판매가")),
            "effectiveFrom": normalize_date(raw.get("적용시작일")),
            "effectiveTo": normalize_date(raw.get("적용종료일")),
            "barcode": str(raw.get("바코드") or "").strip(),
            "isActive": normalize_bool(raw.get("사용상태")),
            "note": str(raw.get("메모") or "").strip(),
        })
    print(json.dumps({"rows": rows}, ensure_ascii=False))


def main():
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)

    export_parser = subparsers.add_parser("export")
    export_parser.add_argument("--input", required=True)
    export_parser.add_argument("--output", required=True)

    import_parser = subparsers.add_parser("import")
    import_parser.add_argument("--input", required=True)

    args = parser.parse_args()
    if args.command == "export":
        export_workbook(args.input, args.output)
    else:
        import_workbook(args.input)


if __name__ == "__main__":
    main()
