#!/usr/bin/env python3
"""픽키도기클럽 계열 월별 세부판매내역(정산서) xlsx 생성.

`종합_(픽키파크)우프_YYYY-MM월_필바이츠 세부판매내역.xlsx` 의 26년 6월 양식을
따른다. 도톤 정산서와 표가 달라 생성기를 나눴다 — 도톤 쪽은 정답지와 맞춰 둔
코드라 공용화하면서 건드릴 이유가 없다.

시트 셋:
  1. 정산서              — 판매내역 종합, 매출계산서 발행 주체별 정산, 메모
  2. 채널별 정산 합계     — 채널 한 블록씩 + 하단 총계
  3. 채널별 판매데이터 정리 — 채널별 상품 단위 내역

CLI: python3 npb_settlement_picky.py --input <spec.json> --output <out.xlsx>
"""

import argparse
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="A6A6A6")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="D9D9D9")
SUB_FILL = PatternFill("solid", fgColor="F2F2F2")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")

MONEY = "#,##0"
PCT = "0.0%"

BOLD = Font(bold=True)
TITLE = Font(size=15, bold=True)
SMALL = Font(size=9, color="808080")

# 채널 블록의 열 구성. 6월 정산서 그대로다.
CHANNEL_HEADERS = [
    "항목", "정가합계", "배송비합계", "할인합계", "최종결제합계",
    "공제(%)", "공제(원)", "정산합계", "정산형태", "비고",
]
DETAIL_HEADERS = [
    "순번", "제품", "판매수량", "정가합계", "배송비", "할인(원)",
    "최종결제합계", "수수료(%)", "수수료(원)", "정산합계",
]


def put(ws, row, col, value, *, font=None, fill=None, fmt=None, align=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align == "left"))
    if border:
        cell.border = BOX
    return cell


def widths(ws, spec):
    for letter, width in spec.items():
        ws.column_dimensions[letter].width = width


def fee_cell(ws, row, col, value):
    """공제(%) 칸. 요율이 품목마다 다른 채널은 숫자 대신 그 사실을 적는다."""
    if isinstance(value, (int, float)):
        return put(ws, row, col, value, fmt=PCT, align="center")
    return put(ws, row, col, value or "", align="center")


# ------------------------------------------------------------------ 정산서
# 확정된 6월 양식과 셀 단위로 같게 그린다. 실비는 이 장에 숫자로 올리지 않는다 —
# 정산에서 공제하지 않는 값을 표에 함께 세우면 공제한 것처럼 읽힌다. 별도 청구
# 사실은 메모 한 줄로만 알린다.
F8 = Font(size=8)
F8B = Font(size=8, bold=True)
F9B = Font(size=9, bold=True)
F10B = Font(size=10, bold=True)
BLUE_FILL = PatternFill("solid", fgColor="FFCAEDFB")   # 소비자정가계·매출계
GREEN_FILL = PatternFill("solid", fgColor="FFDAF2D0")  # 수수료·정산합계
GREY_FILL = PatternFill("solid", fgColor="FFF2F2F2")   # 표 머리
# 우프 정산 줄. 양식은 테마색(accent2 E97132)의 80% 밝은 값을 쓴다 — 테마 참조
# 대신 같은 색을 직접 적는다. 파일을 여는 쪽 테마가 달라도 색이 흔들리지 않는다.
PARTY_FILL = PatternFill("solid", fgColor="FFFBE3D6")


def build_summary(ws, spec):
    widths(ws, {"A": 3.7, "B": 14.5, "C": 16.2, "D": 21.7, "F": 17.5,
                "L": 13.3, "M": 44.5, "N": 8.7, "O": 8.5})
    for row, height in [(2, 24), (3, 13), (4, 16), (6, 23), (11, 16), (12, 28),
                        (15, 10), (16, 15.8)]:
        ws.row_dimensions[row].height = height

    put(ws, 2, 2, spec.get("title") or "월별 세부 판매 현황",
        font=Font(size=11, bold=True), border=False)

    put(ws, 4, 2, "[판매내역 종합]", font=F8B, align="left", border=False)
    put(ws, 4, 7, "vat포함", font=Font(size=9), align="right", border=False)

    for i, (label, font, fill) in enumerate([
        ("계산서 발행 (신고)기준", F8, None),
        ("집계 기준 (채널별 상이)", F8B, None),
        ("소비자정가계", F8B, BLUE_FILL),
        ("매출계", F8B, BLUE_FILL),
        ("수수료 (공제계)", F8B, GREEN_FILL),
        ("정산합계", F8B, GREEN_FILL),
    ]):
        put(ws, 5, 2 + i, label, font=font, fill=fill, align="center")

    rollup = spec.get("rollup") or {}
    put(ws, 6, 2, spec.get("issueBasisDate") or "", font=F10B, align="center")
    put(ws, 6, 3, spec.get("periodRange") or "", font=F9B, align="center")
    for i, key in enumerate(["listTotal", "realSaleTotal", "feeTotal", "settleTotal"]):
        put(ws, 6, 4 + i, rollup.get(key, 0), font=F9B, fmt=MONEY, align="center")

    put(ws, 8, 2, "매출계산서발행", font=F8B, align="left", border=False)
    ws.merge_cells("B8:C8")
    put(ws, 8, 10, "vat포함", font=F8, align="right", border=False)

    for i, label in enumerate(["항목", "A 매출", "B 공제계", "C 정산계", "비고"]):
        put(ws, 9, 2 + i, label, font=F8, fill=GREY_FILL, align="center")
    ws.merge_cells("F9:J9")
    for col in range(6, 11):
        ws.cell(row=9, column=col).fill = GREY_FILL
        ws.cell(row=9, column=col).border = BOX

    row = 10
    put(ws, row, 2, "정산합계", font=F8, align="center")
    put(ws, row, 3, rollup.get("realSaleTotal", 0), font=F8B, fmt=MONEY, align="center")
    put(ws, row, 4, rollup.get("feeTotal", 0), font=F8, fmt=MONEY, align="center")
    put(ws, row, 5, rollup.get("settleTotal", 0), font=F8B, fmt=MONEY, align="center")
    put(ws, row, 6, "", border=True)
    ws.merge_cells("F10:J10")
    for col in range(6, 11):
        ws.cell(row=row, column=col).border = BOX
    row += 1

    notes = {
        "픽키파크": "직접 픽키파크로 업체가 발행 후, 해당 업체에서 픽키파크 계좌로 직접 대금 지급되는 건입니다.",
        "우프": "우프에서 B (공제계) 만큼 수수료 세금계산서를 발행 하고, C 정산계 만큼 대금을 입금해요.  \n"
                "픽키파크에서 A 매출을 기타매출로 신고 하시면 됩니다.",
    }
    # 픽키파크 → 우프 순으로 적는다 (6월 정산서와 같은 순서).
    parties = sorted(
        spec.get("settleParties") or [],
        key=lambda p: 0 if p.get("party") == "픽키파크" else 1,
    )
    for party in parties:
        name = party.get("party") or ""
        fill = PARTY_FILL if name == "우프" else None
        put(ws, row, 2, f"{name} 정산", font=F8, fill=fill, align="center")
        put(ws, row, 3, party.get("sale", 0), font=F8B, fill=fill, fmt=MONEY, align="center")
        put(ws, row, 4, party.get("fee", 0), font=F8, fill=fill, fmt=MONEY, align="center")
        put(ws, row, 5, party.get("settle", 0), font=F8B, fill=fill, fmt=MONEY, align="center")
        cell = put(ws, row, 6, notes.get(name, ""), font=F8)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=10)
        for col in range(6, 11):
            ws.cell(row=row, column=col).border = BOX
        row += 1

    row += 3
    put(ws, row, 2, "[메모/특이사항]", font=F8B, align="left", border=False)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    row += 1
    for line in spec.get("memo") or []:
        put(ws, row, 2, line, font=F8, align="left", border=False)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1


# --------------------------------------------------------- 채널별 정산 합계
def build_channel_totals(ws, spec):
    widths(ws, {"A": 2, "B": 20, "C": 14, "D": 14, "E": 13, "F": 15,
                "G": 12, "H": 13, "I": 14, "J": 11, "K": 46})
    put(ws, 2, 2, f"정산기간 {spec.get('periodRange') or ''}", font=BOLD, border=False)

    row = 4
    for channel in spec.get("channels") or []:
        put(ws, row, 2, channel.get("name") or "", font=BOLD, fill=SUB_FILL)
        basis = channel.get("dateBasis") or ""
        put(ws, row, 3, f"집계기준: {basis}" if basis else "", align="left")
        row += 1
        for i, label in enumerate(CHANNEL_HEADERS):
            put(ws, row, 2 + i, label, font=BOLD, fill=HEAD_FILL, align="center")
        row += 1
        totals = channel.get("totals") or {}
        put(ws, row, 2, "합계", font=BOLD)
        put(ws, row, 3, totals.get("list", 0), fmt=MONEY)
        # 채널이 배송을 맡으면 금액 대신 그 사실을 적는다 (쿠팡배송/자체배송 등).
        ship_note = channel.get("shippingNote")
        if ship_note:
            put(ws, row, 4, ship_note, align="center")
        else:
            put(ws, row, 4, totals.get("shipping", 0), fmt=MONEY)
        put(ws, row, 5, totals.get("discount", 0), fmt=MONEY)
        put(ws, row, 6, totals.get("sale", 0), fmt=MONEY)
        fee_cell(ws, row, 7, channel.get("feeLabel"))
        put(ws, row, 8, totals.get("fee", 0), fmt=MONEY)
        put(ws, row, 9, totals.get("settle", 0), fmt=MONEY)
        put(ws, row, 10, channel.get("settleBy") or "", align="center")
        put(ws, row, 11, channel.get("note") or "", align="left")
        row += 2

    rollup = spec.get("rollup") or {}
    put(ws, row, 11, "VAT 포함", font=SMALL, align="right", border=False)
    row += 1
    summary = [
        ("소비자정가합계", rollup.get("listTotal", 0), "소비자 정상가 기준 판매 합계"),
        ("할인합계", rollup.get("discountTotal", 0),
         "프로모션할인 / 대량구매할인 / 사업자할인 등"),
        ("최종결제액 합계", rollup.get("realSaleTotal", 0),
         "소비자정가합계 - 할인합계 + 배송비 결제분"),
        ("공제합계", rollup.get("feeTotal", 0),
         "거래수수료 / 카드사 등 결제 PG 수수료"),
        ("정산합계", rollup.get("settleTotal", 0),
         "최종결제액 합계 - 공제합계 = 정산합계"),
    ]
    for label, amount, note in summary:
        put(ws, row, 2, label, font=BOLD, fill=SUB_FILL)
        put(ws, row, 3, amount, fmt=MONEY)
        put(ws, row, 4, note, align="left", border=False)
        row += 1
    for party in sorted(spec.get("settleParties") or [],
                        key=lambda p: 0 if p.get("party") == "우프" else 1):
        put(ws, row, 2, f"정산합계({party.get('party')})", font=BOLD, fill=TOTAL_FILL)
        put(ws, row, 3, party.get("settle", 0), fmt=MONEY, fill=TOTAL_FILL)
        row += 1


# ------------------------------------------------- 채널별 판매데이터 정리
def build_channel_detail(ws, spec):
    widths(ws, {"A": 2, "B": 6, "C": 48, "D": 11, "E": 13, "F": 11,
                "G": 12, "H": 14, "I": 11, "J": 13, "K": 14})
    put(ws, 2, 2, f"정산기간 {spec.get('periodRange') or ''}", font=BOLD, border=False)

    row = 4
    for channel in spec.get("channels") or []:
        put(ws, row, 2, channel.get("name") or "", font=BOLD, fill=SUB_FILL)
        basis = channel.get("dateBasis") or ""
        put(ws, row, 3, f"{basis} 기준" if basis else "", align="left")
        row += 1
        for i, label in enumerate(DETAIL_HEADERS):
            put(ws, row, 2 + i, label, font=BOLD, fill=HEAD_FILL, align="center")
        row += 1
        for n, item in enumerate(channel.get("rows") or [], start=1):
            put(ws, row, 2, n, align="center")
            put(ws, row, 3, item.get("label") or "", align="left")
            put(ws, row, 4, item.get("qty", 0), fmt=MONEY)
            put(ws, row, 5, item.get("listTotal", 0), fmt=MONEY)
            put(ws, row, 6, item.get("shipping", 0), fmt=MONEY)
            put(ws, row, 7, item.get("discount", 0), fmt=MONEY)
            put(ws, row, 8, item.get("saleTotal", 0), fmt=MONEY)
            fee_cell(ws, row, 9, item.get("feeRate"))
            put(ws, row, 10, item.get("feeTotal", 0), fmt=MONEY)
            put(ws, row, 11, item.get("settleTotal", 0), fmt=MONEY)
            row += 1
        totals = channel.get("totals") or {}
        put(ws, row, 2, "", fill=TOTAL_FILL)
        put(ws, row, 3, "합계", font=BOLD, fill=TOTAL_FILL)
        for col, key in ((4, "qty"), (5, "list"), (6, "shipping"), (7, "discount"),
                         (8, "sale"), (10, "fee"), (11, "settle")):
            put(ws, row, col, totals.get(key, 0), fmt=MONEY, fill=TOTAL_FILL)
        put(ws, row, 9, "", fill=TOTAL_FILL)
        row += 2


def build_source_sheet(ws, source):
    """올린 채널 파일을 그대로 적는다. 파서는 행을 {열이름: 값} 으로 준다."""
    rows = source.get("rows") or []
    if not rows:
        return
    header, seen = [], set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                header.append(key)
    for c, key in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=key)
        cell.font = BOLD
        cell.fill = SUB_FILL
    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(header, start=1):
            ws.cell(row=r, column=c, value=row.get(key))


def build(spec, path):
    wb = Workbook()
    build_summary(wb.active, spec)
    wb.active.title = "정산서"
    build_channel_totals(wb.create_sheet("채널별 정산 합계"), spec)
    build_channel_detail(wb.create_sheet("채널별 판매데이터 정리"), spec)
    # 올린 채널 파일을 DB) 시트로 뒤에 붙인다. 시트 이름에 못 쓰는 글자는 뺀다.
    used = set(wb.sheetnames)
    for source in spec.get("sources") or []:
        base = f"DB){source.get('label') or source.get('channel') or ''}"
        title = "".join(ch for ch in base if ch not in "\\/*?:[]")[:31] or "DB)"
        name, n = title, 2
        while name in used:
            name = f"{title[:28]}({n})"
            n += 1
        used.add(name)
        build_source_sheet(wb.create_sheet(name), source)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()
    with open(args.input, encoding="utf-8") as fh:
        spec = json.load(fh)
    build(spec, args.output)


if __name__ == "__main__":
    main()
