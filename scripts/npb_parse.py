#!/usr/bin/env python3
"""Parse a monthly per-channel DOTEON sales xlsx into normalized product lines.

Invoked from Node via execFile, mirroring the xlsx_to_json.py pattern:

    python3 npb_parse.py --input <file.xlsx> [--channel <code>]

Emits JSON to stdout:

    {
      "channel": "gongu",
      "sourceFile": "DB_cafe24_영이공구_202605_doteon.xlsx",
      "lines": [
        { "productKey": "spray", "label": "도톤 아웃도어 스프레이 2개세트",
          "qtyEa": 20, "qtyOrders": 10, "eaPerUnit": 2,
          "tier": "2개세트", "raw": { "discountRate": 0.32 } }
      ],
      "warnings": [],
      "meta": {}
    }

Channel is auto-detected from the filename keyword when --channel is omitted.
Unrecognized products are reported in `warnings` (never silently dropped to 0).
Missing/renamed columns degrade to warnings rather than crashes.
"""

import argparse
import json
import os
import re
import sys
import unicodedata

from openpyxl import load_workbook


# --- product identity -------------------------------------------------------

PRODUCT_LABELS = {
    "foot": "도톤 풋클리너",
    "spray": "도톤 아웃도어 스프레이",
}

FOOT_KEYWORDS = ("풋클리너", "발세정", "도톤 풋", "foot")
SPRAY_KEYWORDS = ("아웃도어", "outdoor", "스프레이", "해충방지")

FOOT_BARCODE = "8809879544118"
SPRAY_BARCODE = "8809879544101"

FOOT_CODE_PREFIX = "BT25DTFC"
SPRAY_CODE_PREFIX = "BT25OS"


def norm(v):
    if v is None:
        return ""
    return str(v).strip()


def identify_product(name=None, barcode=None, code=None):
    """Return 'foot' | 'spray' | None from any of name/barcode/상품코드."""
    bc = norm(barcode)
    if bc:
        # barcodes can arrive as int -> str with trailing .0
        bc = bc.split(".")[0]
        if bc == FOOT_BARCODE:
            return "foot"
        if bc == SPRAY_BARCODE:
            return "spray"
    cd = norm(code).upper()
    if cd:
        if cd.startswith(FOOT_CODE_PREFIX):
            return "foot"
        if cd.startswith(SPRAY_CODE_PREFIX):
            return "spray"
    nm = norm(name).lower()
    if nm:
        if any(k.lower() in nm for k in FOOT_KEYWORDS):
            return "foot"
        if any(k.lower() in nm for k in SPRAY_KEYWORDS):
            return "spray"
    return None


# --- channel detection ------------------------------------------------------

# Order matters: 영이공구 must be checked before cafe24 (longer, more specific).
CHANNEL_KEYWORDS = [
    ("영이공구", "gongu"),
    ("cafe24", "cafe24"),
    ("b2b", "b2b"),
    ("대리점", "tailit"),
    ("몽슈슈", "mongshu"),
    ("스마트스토어", "smartstore"),
    ("컬리", "kurly"),
    ("쿠팡", "coupang"),
    ("emart", "molly"),
    ("행사", "terrymarket"),
    ("태리마켓", "terrymarket"),
    ("파마스퀘어", "pharmasquare"),
]


def detect_channel(filename):
    # macOS stores filenames as NFD (decomposed Hangul); normalize to NFC so
    # Korean keyword literals in this file match reliably.
    base = unicodedata.normalize("NFC", os.path.basename(filename))
    for keyword, code in CHANNEL_KEYWORDS:
        if keyword in base:
            return code
    return None


# --- helpers ----------------------------------------------------------------

def to_int(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", "").strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def header_index(header_row):
    """Map stripped header name -> column index."""
    idx = {}
    for i, h in enumerate(header_row):
        key = norm(h)
        if key and key not in idx:
            idx[key] = i
    return idx


def resolve_cols(idx, names, warnings, channel):
    """Return list of resolved indices; append a warning for any missing name."""
    out = []
    for n in names:
        if n in idx:
            out.append(idx[n])
        else:
            out.append(None)
            warnings.append(
                "[%s] expected column '%s' not found in header" % (channel, n)
            )
    return out


class Aggregator:
    """Groups rows into output lines keyed by (productKey, tier)."""

    def __init__(self):
        self._lines = {}

    def add(self, product, orders, ea, ea_per_unit=1, tier=None, raw=None):
        # 미식별 라인은 원본 이름별로 모은다 — 서버가 나중에 매칭한다.
        if product is None:
            tier = (raw or {}).get("sourceName") or tier
        key = (product, tier)
        line = self._lines.get(key)
        if line is None:
            line = {
                "productKey": product,
                "label": PRODUCT_LABELS.get(product, product)
                if product else ((raw or {}).get("sourceName") or "(미식별)"),
                "qtyEa": 0,
                "qtyOrders": 0,
                "eaPerUnit": ea_per_unit,
                "tier": tier,
                "raw": {
                    k: v for k, v in (raw or {}).items() if k not in ("money", "amounts")
                },
                "money": {},
                "amounts": {},
            }
            if tier and product:
                line["label"] = PRODUCT_LABELS.get(product, product) + " " + tier
            self._lines[key] = line
        line["qtyEa"] += ea
        line["qtyOrders"] += orders
        if raw:
            # 금액은 덮어쓰지 않고 더한다. 같은 상품이 여러 행에 걸쳐 오는
            # 것이 정상이고(주문 단위 파일), 덮어쓰면 마지막 한 행만 남는다.
            for src, dest in (("money", line["money"]), ("amounts", line["amounts"])):
                for key, value in (raw.get(src) or {}).items():
                    if value is None:
                        continue
                    dest[key] = (dest.get(key) or 0) + value
            line["raw"].update(
                {k: v for k, v in raw.items() if k not in ("money", "amounts")}
            )

    def lines(self):
        out = []
        for line in self._lines.values():
            item = {
                "productKey": line["productKey"],
                "label": line["label"],
                "qtyEa": line["qtyEa"],
                "qtyOrders": line["qtyOrders"],
                "eaPerUnit": line["eaPerUnit"],
                "raw": line["raw"],
            }
            money = {k: v for k, v in line["money"].items() if v is not None}
            if money:
                item["money"] = money
            if line["amounts"]:
                item["amounts"] = line["amounts"]
            if line["tier"]:
                item["tier"] = line["tier"]
            out.append(item)
        # stable ordering: product then tier
        out.sort(key=lambda x: (x["productKey"] or "~", x.get("tier") or ""))
        return out


def data_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    return rows


def load_sheets(path):
    """Read every sheet as (name, rows). Supports .xlsx/.xlsm via openpyxl and
    .csv via the csv module (many channel exports are CSV). Raises on unreadable
    input so main() can turn it into a visible warning.

    거래처가 보내는 파일은 시트 하나라는 보장이 없다 — 요약 시트 뒤에 상세가
    붙어 오거나, 원본 다운로드가 두 번째 시트에 있는 경우가 흔하다."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".txt"):
        import csv
        with open(path, newline="", encoding="utf-8-sig") as fh:
            return [("csv", [tuple(row) for row in csv.reader(fh)])]
    if ext == ".xls":
        raise ValueError(
            "구형 .xls 형식은 지원하지 않습니다. .xlsx 또는 .csv로 저장 후 올려주세요."
        )
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = []
    for ws in wb.worksheets:
        # 거래처 파일은 <dimension> 이 틀리게 적혀 오는 경우가 흔하다. read_only
        # 모드는 그 값을 그대로 믿기 때문에 195행짜리 시트가 1행으로 읽힌다
        # (네이버 결제정산 파일이 실제로 그랬다). 실제 셀을 세어 다시 잡는다.
        try:
            ws.reset_dimensions()
        except AttributeError:
            pass
        sheets.append((ws.title, list(ws.iter_rows(values_only=True))))
    return sheets


def load_rows(path):
    """First sheet only — the fixed-recipe DOTEON parsers expect this shape."""
    sheets = load_sheets(path)
    return sheets[0][1] if sheets else []


# --- per-channel parsers ----------------------------------------------------

def parse_cafe24(rows, channel, warnings, is_gongu=False):
    """cafe24 family (cafe24, gongu, b2b, tailit): J=상품명, K=옵션, L=수량."""
    agg = Aggregator()
    if not rows:
        warnings.append("[%s] empty sheet" % channel)
        return agg.lines()
    idx = header_index(rows[0])
    ci_name, ci_opt, ci_qty, ci_refund = resolve_cols(
        idx,
        ["주문상품명(기본)", "상품옵션(기본)", "수량", "환불완료일"],
        warnings,
        channel,
    )
    if ci_name is None or ci_qty is None:
        return agg.lines()
    for r in rows[1:]:
        if r is None or all(x is None for x in r):
            continue
        # skip refunded rows
        if ci_refund is not None and ci_refund < len(r) and norm(r[ci_refund]):
            continue
        name = r[ci_name] if ci_name < len(r) else None
        qty = to_int(r[ci_qty]) if ci_qty < len(r) else 0
        product = identify_product(name=name)
        if product is None:
            warnings.append(
                "[%s] unrecognized product (qty=%d): %r" % (channel, qty, norm(name))
            )
            continue
        if is_gongu:
            opt = norm(r[ci_opt]) if (ci_opt is not None and ci_opt < len(r)) else ""
            m_ea = re.search(r"(\d+)\s*개", opt)
            m_disc = re.search(r"\((\d+)%할인\)", opt)
            ea_per_unit = int(m_ea.group(1)) if m_ea else 1
            disc = int(m_disc.group(1)) / 100.0 if m_disc else None
            if ea_per_unit == 1:
                tier = "1개"
            else:
                tier = "%d개세트" % ea_per_unit
            if not m_ea:
                warnings.append(
                    "[%s] could not parse 공구 tier from option: %r" % (channel, opt)
                )
            agg.add(
                product,
                orders=qty,  # 수량 is always 1 per 공구 row
                ea=qty * ea_per_unit,
                ea_per_unit=ea_per_unit,
                tier=tier,
                raw={"discountRate": disc, "option": opt},
            )
        else:
            agg.add(product, orders=1, ea=qty, ea_per_unit=1)
    return agg.lines()


def parse_mongshu(rows, channel, warnings):
    """몽슈슈 재고표: product col B, code col C, qty from 판매 col H."""
    agg = Aggregator()
    B, C, H = 1, 2, 7  # 제품, 상품코드, 판매
    for r in rows:
        if r is None or len(r) <= H:
            continue
        seq = r[0] if len(r) > 0 else None
        # data rows carry a numeric 순번 in col A
        if not isinstance(seq, (int, float)):
            continue
        name = r[B]
        code = r[C]
        qty = to_int(r[H])
        product = identify_product(name=name, code=code)
        if product is None:
            warnings.append(
                "[%s] non-doteon / unrecognized product (판매=%d): %r / %r"
                % (channel, qty, norm(name), norm(code))
            )
            continue
        agg.add(product, orders=1, ea=qty, ea_per_unit=1)
    return agg.lines()


def parse_smartstore(rows, channel, warnings):
    """Naver smartstore: 상품명 col O, 수량 col R, 판매채널 col D."""
    agg = Aggregator()
    if not rows:
        warnings.append("[%s] empty sheet" % channel)
        return agg.lines()
    idx = header_index(rows[0])
    ci_name, ci_qty = resolve_cols(idx, ["상품명", "수량"], warnings, channel)
    if ci_name is None or ci_qty is None:
        return agg.lines()
    for r in rows[1:]:
        if r is None or all(x is None for x in r):
            continue
        name = r[ci_name] if ci_name < len(r) else None
        qty = to_int(r[ci_qty]) if ci_qty < len(r) else 0
        product = identify_product(name=name)
        if product is None:
            warnings.append(
                "[%s] unrecognized product (qty=%d): %r" % (channel, qty, norm(name))
            )
            continue
        agg.add(product, orders=1, ea=qty, ea_per_unit=1)
    return agg.lines()


def parse_kurly(rows, channel, warnings):
    """컬리: 상품명 col F / 옵션명 col G, 수량 col M."""
    agg = Aggregator()
    if not rows:
        warnings.append("[%s] empty sheet" % channel)
        return agg.lines()
    idx = header_index(rows[0])
    ci_name, ci_opt, ci_qty = resolve_cols(
        idx, ["상품명", "옵션명", "수량"], warnings, channel
    )
    if ci_name is None or ci_qty is None:
        return agg.lines()
    for r in rows[1:]:
        if r is None or all(x is None for x in r):
            continue
        name = r[ci_name] if ci_name < len(r) else None
        opt = r[ci_opt] if (ci_opt is not None and ci_opt < len(r)) else None
        qty = to_int(r[ci_qty]) if ci_qty < len(r) else 0
        product = identify_product(name=name) or identify_product(name=opt)
        if product is None:
            warnings.append(
                "[%s] unrecognized product (qty=%d): %r" % (channel, qty, norm(name))
            )
            continue
        agg.add(product, orders=1, ea=qty, ea_per_unit=1)
    return agg.lines()


def parse_coupang(rows, channel, warnings):
    """쿠팡: SKU명 col D, 수량 col H, rows where 구분='발주'. Bundle-aware."""
    agg = Aggregator()
    if not rows:
        warnings.append("[%s] empty sheet" % channel)
        return agg.lines()
    idx = header_index(rows[0])
    ci_gubun, ci_sku, ci_qty = resolve_cols(
        idx, ["구분", "SKU명", "수량"], warnings, channel
    )
    if ci_sku is None or ci_qty is None:
        return agg.lines()
    distinct_skus = []
    for r in rows[1:]:
        if r is None or all(x is None for x in r):
            continue
        if ci_gubun is not None and ci_gubun < len(r):
            if norm(r[ci_gubun]) != "발주":
                continue
        sku = norm(r[ci_sku]) if ci_sku < len(r) else ""
        qty = to_int(r[ci_qty]) if ci_qty < len(r) else 0
        product = identify_product(name=sku)
        if product is None:
            warnings.append(
                "[%s] unrecognized SKU (수량=%d): %r" % (channel, qty, sku)
            )
            continue
        if sku not in distinct_skus:
            distinct_skus.append(sku)
        # bundle size from SKU name (e.g. '2개', '3개세트'); default 1
        m_ea = re.search(r"(\d+)\s*개", sku)
        ea_per_unit = int(m_ea.group(1)) if m_ea else 1
        tier = None if ea_per_unit == 1 else "%d개" % ea_per_unit
        agg.add(
            product,
            orders=qty,
            ea=qty * ea_per_unit,
            ea_per_unit=ea_per_unit,
            tier=tier,
            raw={"sku": sku},
        )
    if distinct_skus:
        warnings.append(
            "[%s] distinct SKU names (verify bundle composition): %s"
            % (channel, "; ".join(distinct_skus))
        )
    return agg.lines()


def parse_emart(rows, channel, warnings, meta):
    """emart(몰리스): 상품코드(barcode) col G / 상품명 col H, 납품량 col N. VAT별도."""
    agg = Aggregator()
    if not rows:
        warnings.append("[%s] empty sheet" % channel)
        return agg.lines()
    idx = header_index(rows[0])
    ci_code, ci_name, ci_qty, ci_amt = resolve_cols(
        idx, ["상품코드", "상품명", "납품량", "납품금액"], warnings, channel
    )
    if ci_qty is None or (ci_code is None and ci_name is None):
        return agg.lines()
    meta["vatIncluded"] = False
    meta["basis"] = "납품금액"
    warnings.append("[%s] VAT 별도 (unit price excludes VAT)" % channel)
    for r in rows[1:]:
        if r is None or all(x is None for x in r):
            continue
        barcode = r[ci_code] if (ci_code is not None and ci_code < len(r)) else None
        name = r[ci_name] if (ci_name is not None and ci_name < len(r)) else None
        qty = to_int(r[ci_qty]) if ci_qty < len(r) else 0
        amt = to_int(r[ci_amt]) if (ci_amt is not None and ci_amt < len(r)) else 0
        product = identify_product(name=name, barcode=barcode)
        if product is None:
            warnings.append(
                "[%s] unrecognized product (납품량=%d): %r" % (channel, qty, norm(name))
            )
            continue
        unit = round(amt / qty) if qty else None
        agg.add(
            product,
            orders=1,
            ea=qty,
            ea_per_unit=1,
            raw={"vatIncluded": False, "basis": "납품금액", "unitPrice": unit},
        )
    return agg.lines()


def parse_pharmasquare(rows, channel, warnings):
    """파마스퀘어(대리점형 45%): 바코드 col C, 판매수량 col E."""
    agg = Aggregator()
    if not rows:
        warnings.append("[%s] empty sheet" % channel)
        return agg.lines()
    idx = header_index(rows[0])
    ci_code, ci_name, ci_qty = resolve_cols(
        idx, ["바코드", "상품", "판매수량"], warnings, channel
    )
    if ci_qty is None or ci_code is None:
        return agg.lines()
    for r in rows[1:]:
        if r is None or all(x is None for x in r):
            continue
        barcode = r[ci_code] if ci_code < len(r) else None
        # subtotal rows have no barcode -> skip
        if not norm(barcode):
            continue
        name = r[ci_name] if (ci_name is not None and ci_name < len(r)) else None
        qty = to_int(r[ci_qty]) if ci_qty < len(r) else 0
        product = identify_product(barcode=barcode, name=name)
        if product is None:
            warnings.append(
                "[%s] unrecognized product (판매수량=%d): %r" % (channel, qty, norm(name))
            )
            continue
        agg.add(
            product, orders=1, ea=qty, ea_per_unit=1, raw={"feeRate": 0.45}
        )
    return agg.lines()


def parse_terrymarket(rows, channel, warnings):
    """태리마켓/행사: 품목 col A, 판매수량 col B, 마켓 할인가 col C, 행사명 col H."""
    agg = Aggregator()
    if not rows:
        warnings.append("[%s] empty sheet" % channel)
        return agg.lines()
    idx = header_index(rows[0])
    ci_item, ci_qty, ci_price, ci_event = resolve_cols(
        idx, ["품목", "판매수량", "마켓 할인가", "행사명"], warnings, channel
    )
    if ci_item is None or ci_qty is None:
        return agg.lines()
    for r in rows[1:]:
        if r is None or all(x is None for x in r):
            continue
        item = norm(r[ci_item]) if ci_item < len(r) else ""
        if not item or "합계" in item:
            continue
        qty = to_int(r[ci_qty]) if ci_qty < len(r) else 0
        price = (
            to_int(r[ci_price]) if (ci_price is not None and ci_price < len(r)) else None
        )
        event = (
            norm(r[ci_event]) if (ci_event is not None and ci_event < len(r)) else ""
        )
        product = identify_product(name=item)
        if product is None:
            warnings.append(
                "[%s] unrecognized 품목 (판매수량=%d): %r" % (channel, qty, item)
            )
            continue
        agg.add(
            product,
            orders=1,
            ea=qty,
            ea_per_unit=1,
            raw={"marketPrice": price, "eventName": event},
        )
    return agg.lines()


# --- dispatch ---------------------------------------------------------------

# --- format-tolerant scanner ------------------------------------------------
#
# 거래처가 보내는 정산자료는 같은 채널이라도 매달 양식이 달라진다. 헤더가 1행에
# 있다는 보장도, 시트가 하나라는 보장도, 컬럼 이름이 그대로라는 보장도 없다.
# 그래서 "N행 M열" 을 고정하지 않고 다음 순서로 읽는다:
#
#   1. 모든 시트의 모든 행을 훑어 '헤더처럼 보이는 행'을 전부 찾는다
#      (수량 계열 컬럼 + 상품 식별 컬럼이 같은 행에 있으면 헤더로 본다)
#   2. 각 헤더 아래 데이터 구간을 잘라 실제로 읽히는 행 수로 점수를 매긴다
#   3. 가장 잘 읽히는 블록 하나를 채택하고, 나머지 후보도 함께 돌려준다
#      — 잘못 골랐을 때 화면에서 사람이 블록을 바꿀 수 있어야 하기 때문이다
#
# 스파크펫 파일처럼 한 시트에 요약 블록과 상세 블록이 같이 있는 경우, 둘 다
# 집계하면 매출이 두 배가 된다. 그래서 여러 블록을 합치지 않고 하나만 고른다.

# 역할별 컬럼 이름. exact 는 헤더가 정확히 일치할 때(가장 확실), substr 은
# 부분 일치 + 제외어. 제외어가 없으면 "배송비 합계"가 매출 합계로 잡히는 식의
# 사고가 난다.
ROLE_EXACT = {
    "qty": ["수량", "판매수량", "주문수량", "결제수량", "구매수량", "출고수량", "발주수량",
            "판매 수량", "정산 대상 수량", "정산대상수량", "판매", "납품량", "상품 수량"],
    "name": ["상품명", "sku명", "품목명", "제품명", "품목", "상품", "제품",
             "주문상품명", "주문상품명(기본)", "주문상품명(옵션포함)", "통합상품명(출력용)",
             "상품 이름", "재고매칭(1)상품명", "판매처상품명"],
    "option": ["옵션명", "옵션정보", "상품옵션", "옵션내용", "옵션", "상품옵션(기본)",
               "옵션 이름", "옵션값"],
    "barcode": ["바코드", "바코드번호", "barcode", "재고매칭(1)바코드번호"],
    "code": ["상품코드", "sku번호", "품목관리코드", "판매자상품코드", "판매자 상품코드",
             "자체상품코드", "품목코드", "sku", "딜코드"],
    "listUnit": ["정가", "소비자가", "권장소비자가", "정상가", "판매정가"],
    "listAmount": ["정가합계", "정가계", "정가합", "정상가합계", "소비자가합계"],
    "saleUnit": ["판매가", "상품가격", "적용단가", "단가", "공급가", "총단가", "공급가(소매)"],
    "saleAmount": ["합계", "힙계", "총합계", "판매합계", "총판매가", "주문금액", "총주문금액",
                   "총 주문금액", "최종 상품별 총 주문금액", "최종결제합계", "총결제액",
                   "공급가합계", "공급가합", "총공급가액", "판매 합계 금액",
                   "정산기준금액(a)", "납품금액", "매출"],
    "discount": ["할인금액", "할인액", "할인가", "할인(원)", "최종 상품별 할인액",
                 "판매자 부담 할인액", "할인"],
    "shipping": ["배송비", "배송비합계", "배송비 합계", "고객배송비", "배송비계"],
    "fee": ["수수료", "수수료(원)", "위탁수수료", "npay 수수료(b)", "매출연동 수수료 합계(c)",
            "네이버페이 주문관리 수수료", "매출연동 수수료", "무이자할부 수수료(d)",
            "공제(원)", "공제"],
    "settle": ["정산금액", "정산예정금액", "정산합계", "정산 합계 금액", "정산금",
               "베럴즈 정산금"],
    "gubun": ["구분", "주문상태", "판매구분", "상태", "재고작업구분", "청구구분"],
    "date": ["일자", "거래일", "주문일자", "주문일시", "구매확정일", "배송완료일",
             "정산기준일", "입출고일자", "입고/반출시각", "발송일시", "세금신고기준일"],
}

# (포함어, 제외어)
ROLE_SUBSTR = {
    "qty": (["수량"], ["반품", "취소", "환불", "재고", "입고", "반송", "잔여", "누적",
                       "박스", "추가", "묶음"]),
    "name": (["상품명", "품목명", "제품명", "sku명"], ["코드", "번호", "분류", "구분"]),
    "option": (["옵션"], ["번호", "코드", "가격"]),
    "barcode": (["바코드"], []),
    "code": (["상품코드", "품목코드", "관리코드"], []),
    "listUnit": (["정가", "소비자가"], ["합", "계"]),
    "listAmount": (["정가합", "정가계"], []),
    "saleUnit": (["판매가", "공급가", "단가"], ["합", "계", "총"]),
    "saleAmount": (["금액", "합계", "매출"], ["배송", "수수료", "할인", "반품", "취소",
                                              "세액", "부가", "공제", "정산", "원가", "입고"]),
    "discount": (["할인"], ["율", "%", "형태", "유형"]),
    "shipping": (["배송비"], ["묶음", "형태", "유형", "할인"]),
    "fee": (["수수료"], ["율", "%"]),
    "settle": (["정산금", "정산합계"], ["율", "%", "일", "상태"]),
    "gubun": (["구분", "주문상태"], []),
    "date": (["일자", "확정일", "완료일", "기준일", "일시"], []),
}

# 이 역할들은 컬럼이 여러 개면 전부 더한다. 네이버는 수수료를 Npay/매출연동/
# 무이자할부 세 컬럼으로 쪼개 주므로 하나만 집으면 공제가 모자란다.
SUMMED_ROLES = {"fee"}

# 이 값이 들어 있는 구분/상태 행은 매출이 아니다.
CANCEL_WORDS = ("취소", "반품", "환불", "반출", "미결제", "결제대기", "입금대기", "폐기")

# 합계·소계 행은 데이터가 아니다.
TOTAL_WORDS = ("합계", "소계", "총계", "총합", "누계")


def header_role_map(row):
    """한 행을 헤더로 보고 역할 -> 컬럼 인덱스(들) 을 만든다."""
    roles = {}
    labels = {}
    cells = []
    for i, raw in enumerate(row or []):
        text = norm(raw)
        if not text or len(text) > 40:
            cells.append(None)
            continue
        cells.append((text, text.lower().replace(" ", "")))
    # 1차: 정확 일치
    for i, cell in enumerate(cells):
        if not cell:
            continue
        text, key = cell
        for role, names in ROLE_EXACT.items():
            if key in [n.replace(" ", "") for n in names]:
                roles.setdefault(role, []).append(i)
                labels.setdefault(role, []).append(text)
                break
    # 2차: 부분 일치 (정확 일치로 이미 잡힌 컬럼은 건너뛴다)
    taken = {i for idxs in roles.values() for i in idxs}
    for i, cell in enumerate(cells):
        if not cell or i in taken:
            continue
        text, key = cell
        for role, (includes, excludes) in ROLE_SUBSTR.items():
            if any(x in key for x in excludes):
                continue
            if any(inc.replace(" ", "") in key for inc in includes):
                roles.setdefault(role, []).append(i)
                labels.setdefault(role, []).append(text)
                break
    return roles, labels


def to_num(v):
    """금액용. 음수(수수료 -976)와 소수점을 잃지 않는다."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("₩", "").strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def cell_of(row, i):
    return row[i] if (i is not None and i < len(row)) else None


def role_value(row, roles, role):
    """역할에 해당하는 값. SUMMED_ROLES 는 컬럼을 모두 더한다."""
    idxs = roles.get(role) or []
    if not idxs:
        return None
    if role in SUMMED_ROLES:
        total = 0.0
        seen = False
        for i in idxs:
            n = to_num(cell_of(row, i))
            if n is None:
                continue
            total += abs(n)
            seen = True
        return total if seen else None
    return to_num(cell_of(row, idxs[0]))


def role_text(row, roles, role):
    idxs = roles.get(role) or []
    for i in idxs:
        text = norm(cell_of(row, i))
        if text:
            return text
    return ""


MONEY_ROLES = ("listUnit", "listAmount", "saleUnit", "saleAmount", "discount",
               "shipping", "fee", "settle")


def name_fallback_col(rows, start, end, roles):
    """상품명 컬럼에 머리글이 없는 손수 만든 표를 위한 보정.

    철수마켓 정산서처럼 품목 열의 머리글 칸이 비어 있는 파일이 흔하다. 이때는
    데이터 구간에서 글자가 가장 많이 들어 있는 열을 상품명으로 본다."""
    taken = {i for role, idxs in roles.items() if role != "gubun" for i in idxs}
    counts = {}
    for row in rows[start:end]:
        if row is None:
            continue
        for i, value in enumerate(row):
            if i in taken:
                continue
            text = norm(value)
            if not text or to_num(value) is not None:
                continue
            counts[i] = counts.get(i, 0) + 1
    if not counts:
        return None
    best = max(counts.items(), key=lambda kv: (kv[1], -kv[0]))
    return best[0] if best[1] >= 2 else None


def numeric_columns(rows, start, end, header_row):
    """데이터 구간에서 숫자가 들어 있는 열 전부 -> {인덱스: 머리글}.

    자동 추정이 틀렸을 때 화면에서 다른 열을 고를 수 있어야 하므로, 역할을
    못 붙인 금액 열도 이름과 함께 그대로 넘긴다."""
    out = {}
    for i, raw in enumerate(header_row or []):
        label = norm(raw)
        if label:
            out[i] = label
    found = {}
    for row in rows[start:end]:
        if row is None:
            continue
        for i, value in enumerate(row):
            if to_num(value) is None:
                continue
            found[i] = found.get(i, 0) + 1
    return {i: out.get(i) or ("열%d" % (i + 1)) for i in found if found[i] >= 2}


def find_blocks(sheets):
    """모든 시트에서 헤더 후보와 그 아래 데이터 구간을 찾는다."""
    blocks = []
    for sheet_name, rows in sheets:
        headers = []
        for i, row in enumerate(rows):
            if row is None:
                continue
            # 머리글 행에는 숫자가 없다. 이 조건이 없으면 "반납수량 27+1(DP)"
            # 같은 비고 칸 때문에 데이터 행이 머리글로 잡혀 표가 잘게 쪼개진다.
            if any(to_num(v) is not None for v in row):
                continue
            roles, labels = header_role_map(row)
            # 역할이 하나뿐인 행은 우연히 맞은 것으로 본다.
            if len(roles) < 2:
                continue
            # 수량이든 금액이든 집계할 값이 있어야 정산에 쓸 수 있는 블록이다.
            # 네이버 결제정산 파일처럼 수량이 아예 없고 금액만 있는 양식도 받아야
            # 한다. 상품 열은 아래에서 머리글 없는 경우까지 찾아본다.
            if "qty" not in roles and not any(r in roles for r in MONEY_ROLES):
                continue
            headers.append((i, roles, labels))
        for pos, (i, roles, labels) in enumerate(headers):
            end = headers[pos + 1][0] if pos + 1 < len(headers) else len(rows)
            if "name" not in roles:
                # 손수 만든 정산표는 품목 열의 머리글 칸이 비어 있곤 한다.
                fallback = name_fallback_col(rows, i + 1, end, roles)
                if fallback is not None:
                    roles = dict(roles)
                    labels = dict(labels)
                    roles["name"] = [fallback]
                    labels["name"] = [norm(cell_of(rows[i], fallback)) or "(머리글 없음)"]
            if not any(r in roles for r in ("name", "code", "barcode", "gubun")):
                continue
            blocks.append({
                "sheet": sheet_name,
                "headerRow": i,
                "roles": roles,
                "labels": labels,
                "start": i + 1,
                "end": end,
                "rows": rows,
                "numeric": numeric_columns(rows, i + 1, end, rows[i]),
            })
    return blocks


# 구분 칸이 이 값이면 상품 행이 아니라 부대비용 행이다. 네이버 결제정산은
# 배송비 행에도 상품명이 남아 있는 경우가 있어, 이름을 먼저 보면 배송비가
# 상품 매출로 합산된다(6월 자료에서 3건, 9,000원).
CATEGORY_GUBUN = ("배송비", "수수료", "할인", "쿠폰", "포인트", "적립", "부담금")


def block_row_key(row, roles):
    """행에서 상품을 가리키는 이름. 없으면 구분(배송비 등)이라도 쓴다."""
    gubun = role_text(row, roles, "gubun")
    if gubun and any(gubun.startswith(w) or gubun == w for w in CATEGORY_GUBUN):
        return gubun
    name = role_text(row, roles, "name")
    if name:
        return name
    option = role_text(row, roles, "option")
    if option and option not in ("단일상품", "-"):
        return option
    return role_text(row, roles, "gubun")


def score_block(block):
    """실제로 한 줄로 읽히는 행 수. 이 점수로 블록을 고른다."""
    hits = 0
    for row in block["rows"][block["start"]:block["end"]]:
        if row is None or all(x is None for x in row):
            continue
        name = block_row_key(row, block["roles"])
        code = role_text(row, block["roles"], "code") or role_text(row, block["roles"], "barcode")
        if not name and not code:
            continue
        if any(w in name for w in TOTAL_WORDS):
            continue
        qty = role_value(row, block["roles"], "qty")
        if qty is not None and qty > 0:
            hits += 1
            continue
        if any(role_value(row, block["roles"], r) for r in MONEY_ROLES):
            hits += 1
    return hits


# 주문 단위로 같은 값이 반복되는 열은 라인 금액으로 쓰면 안 된다. 카페24의
# "총 주문금액" 은 한 주문의 모든 품목 행에 같은 값이 들어 있어, 그대로 더하면
# 주문에 품목이 셋이면 매출이 세 배가 된다.
ORDER_LEVEL_HINTS = ("총주문금액", "총결제금액", "주문금액합", "결제금액")
LINE_LEVEL_HINTS = ("상품별", "품목별", "라인")


def looks_order_level(label):
    key = norm(label).lower().replace(" ", "")
    if any(h in key for h in LINE_LEVEL_HINTS):
        return False
    return any(h in key for h in ORDER_LEVEL_HINTS)


def scan_rows(block, channel, warnings):
    """채택한 블록을 상품 라인으로 집계한다. 금액이 있으면 함께 담는다."""
    agg = Aggregator()
    roles = block["roles"]
    numeric = block.get("numeric") or {}
    skipped_cancel = 0

    # 자동 추정이 주문 단위 열을 매출로 집었으면 물린다 — 합계가 배로 튄다.
    sale_idxs = roles.get("saleAmount") or []
    if sale_idxs and looks_order_level((block["labels"].get("saleAmount") or [""])[0]):
        warnings.append(
            "[%s] '%s' 는 주문 단위 합계로 보여 매출에서 제외했습니다. 필요하면 화면에서 지정하세요."
            % (channel, block["labels"]["saleAmount"][0])
        )
        roles = {k: v for k, v in roles.items() if k != "saleAmount"}

    for row in block["rows"][block["start"]:block["end"]]:
        if row is None or all(x is None for x in row):
            continue
        name = block_row_key(row, roles)
        option = role_text(row, roles, "option")
        code = role_text(row, roles, "code")
        barcode = role_text(row, roles, "barcode")
        if not name and not code and not barcode:
            continue
        if any(w in name for w in TOTAL_WORDS):
            continue
        gubun = role_text(row, roles, "gubun")
        if gubun and any(w in gubun for w in CANCEL_WORDS):
            skipped_cancel += 1
            continue
        qty_raw = role_value(row, roles, "qty")
        qty = int(round(qty_raw)) if qty_raw is not None else 0
        amounts = {}
        for i, label in numeric.items():
            value = to_num(cell_of(row, i))
            if value is not None:
                amounts[label] = amounts.get(label, 0.0) + value
        # 수량도 금액도 없으면 빈 행이다.
        if qty <= 0 and not amounts:
            continue

        list_unit = role_value(row, roles, "listUnit")
        list_amount = role_value(row, roles, "listAmount")
        sale_unit = role_value(row, roles, "saleUnit")
        sale_amount = role_value(row, roles, "saleAmount")
        discount = role_value(row, roles, "discount")
        shipping = role_value(row, roles, "shipping")
        fee = role_value(row, roles, "fee")
        settle = role_value(row, roles, "settle")

        # 단가만 있으면 수량을 곱해 라인 합계를 만든다. 합계 열이 이미 있으면
        # 그쪽이 거래처가 확정한 금액이므로 건드리지 않는다.
        if qty > 0:
            if list_amount is None and list_unit is not None:
                list_amount = list_unit * qty
            if sale_amount is None and sale_unit is not None:
                sale_amount = sale_unit * qty

        # 화면에 뜨는 이름. 옵션이 다르면 다른 상품이므로 함께 붙인다 —
        # "필바이츠 45g (수량=3개)" 와 "(수량=1개)" 는 단가가 다르다.
        source = name or code or barcode
        if option and option not in ("단일상품", "-") and option != name:
            source = "%s (%s)" % (source, option)

        # 상품 판정은 서버가 브랜드 상품표·별칭으로 한다. 여기서 판정하면
        # 파이썬에 박힌 도톤 키워드가 다른 브랜드 파일까지 잡아 버린다.
        agg.add(
            None,
            orders=1,
            ea=qty,
            ea_per_unit=1,
            raw={
                "sourceName": source,
                "productName": name,
                "option": option,
                "code": code,
                "barcode": barcode,
                "money": {
                    "list": list_amount,
                    "sale": sale_amount,
                    "discount": discount,
                    "shipping": shipping,
                    "fee": fee,
                    "settle": settle,
                },
                "amounts": amounts,
            },
        )
    if skipped_cancel:
        warnings.append(
            "[%s] 취소/반품/반출로 보이는 %d행을 제외했습니다." % (channel, skipped_cancel)
        )
    return agg.lines()


def parse_scan(sheets, channel, warnings, meta, hint=None):
    """전용 레시피가 없는 모든 채널의 기본 경로."""
    blocks = find_blocks(sheets)
    if not blocks:
        seen = []
        for sheet_name, rows in sheets[:3]:
            for row in (rows or [])[:8]:
                names = [norm(v) for v in (row or []) if norm(v)]
                if len(names) >= 3:
                    seen = names[:12]
                    break
            if seen:
                break
        warnings.append(
            "[%s] 상품·수량 컬럼이 있는 표를 찾지 못했습니다. 읽은 머리글: %s"
            % (channel, ", ".join(seen) or "(없음)")
        )
        meta["blocks"] = []
        return []

    scored = []
    for b in blocks:
        scored.append((score_block(b), b))
    scored.sort(key=lambda x: -x[0])

    # 화면에서 블록을 바꿀 수 있도록 후보를 전부 넘긴다.
    meta["blocks"] = [
        {
            "sheet": b["sheet"],
            "headerRow": b["headerRow"] + 1,
            "rowCount": score,
            "columns": {role: labels[0] for role, labels in b["labels"].items()},
        }
        for score, b in scored[:8]
    ]

    chosen = None
    if hint:
        for score, b in scored:
            if b["sheet"] == hint.get("sheet") and b["headerRow"] + 1 == hint.get("headerRow"):
                chosen = b
                break
        if chosen is None:
            warnings.append("[%s] 지정한 표를 찾지 못해 자동 선택으로 돌아갑니다." % channel)
    if chosen is None:
        chosen = scored[0][1]
    if scored[0][0] <= 0:
        warnings.append("[%s] 표는 찾았지만 수량이 있는 행이 없습니다." % channel)

    meta["chosenBlock"] = {
        "sheet": chosen["sheet"],
        "headerRow": chosen["headerRow"] + 1,
        "columns": {role: labels[0] for role, labels in chosen["labels"].items()},
        # 자동 추정이 틀렸을 때 고를 수 있는 금액 열 목록.
        "numericColumns": sorted(set((chosen.get("numeric") or {}).values())),
    }
    if len(scored) > 1:
        warnings.append(
            "[%s] '%s' 시트 %d행 머리글로 읽었습니다 (후보 %d개 중). 다르면 화면에서 바꾸세요."
            % (channel, chosen["sheet"], chosen["headerRow"] + 1, len(scored))
        )
    return scan_rows(chosen, channel, warnings)


def parse_generic(rows, channel, warnings):
    """예전 진입점 — 한 시트만 받는 호출자를 위해 남겨 둔다."""
    return parse_scan([("sheet1", rows or [])], channel, warnings, {})


def parse(channel, sheets, warnings, meta, hint=None):
    rows = sheets[0][1] if sheets else []
    if channel == "gongu":
        return parse_cafe24(rows, channel, warnings, is_gongu=True)
    if channel in ("cafe24", "b2b", "tailit"):
        return parse_cafe24(rows, channel, warnings, is_gongu=False)
    if channel == "mongshu":
        return parse_mongshu(rows, channel, warnings)
    if channel == "smartstore":
        return parse_smartstore(rows, channel, warnings)
    if channel == "kurly":
        return parse_kurly(rows, channel, warnings)
    if channel == "coupang":
        return parse_coupang(rows, channel, warnings)
    if channel == "molly":
        return parse_emart(rows, channel, warnings, meta)
    if channel == "pharmasquare":
        return parse_pharmasquare(rows, channel, warnings)
    if channel == "terrymarket":
        return parse_terrymarket(rows, channel, warnings)
    # 전용 레시피가 없으면 스캐너로 넘긴다. 화면에서 채널을 추가하고 바로
    # 파일을 올릴 수 있어야 한다.
    return parse_scan(sheets, channel, warnings, meta, hint)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--channel", default="")
    # 사람이 화면에서 "이 시트의 이 행이 머리글" 이라고 고쳐 준 경우.
    ap.add_argument("--block-sheet", default="")
    ap.add_argument("--block-header-row", type=int, default=0)
    args = ap.parse_args()

    warnings = []
    meta = {}
    detected = detect_channel(args.input)
    channel = args.channel or detected
    if not channel:
        # 채널을 몰라도 표는 읽는다. 어느 채널인지는 화면에서 고르면 된다 —
        # 파일명 규칙에 안 맞는다고 파일을 통째로 버릴 이유가 없다.
        warnings.append(
            "파일명으로 채널을 알 수 없습니다 (%s). 화면에서 채널을 지정하세요."
            % os.path.basename(args.input)
        )
        channel = "unknown"

    hint = None
    if args.block_sheet and args.block_header_row:
        hint = {"sheet": args.block_sheet, "headerRow": args.block_header_row}

    try:
        sheets = load_sheets(args.input)
    except Exception as exc:  # unreadable file -> visible warning, not a crash
        warnings.append(
            "파일을 읽을 수 없습니다 (%s): %s"
            % (os.path.basename(args.input), exc)
        )
        sheets = None

    lines = parse(channel, sheets, warnings, meta, hint) if sheets else []

    result = {
        "channel": channel,
        "detectedChannel": detected,
        "sourceFile": os.path.basename(args.input),
        "lines": lines,
        "warnings": warnings,
        "meta": meta,
    }
    json.dump(result, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
