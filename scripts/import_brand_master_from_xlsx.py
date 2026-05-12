from __future__ import annotations

import json
import re
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DB_PATH = DATA_DIR / "db.json"
WORKBOOK_PATH = DATA_DIR / "2026_선매입 브랜드 관리대장.xlsx"


BANK_NAMES = [
    "국민은행",
    "신한은행",
    "기업은행",
    "우리은행",
    "하나은행",
    "농협은행",
    "카카오뱅크",
    "토스뱅크",
    "SC제일은행",
    "국민",
    "신한",
    "기업",
    "우리",
    "하나",
    "농협",
    "카카오",
    "토스",
    "SC",
]


def normalize_name(value: str) -> str:
    text = re.sub(r"^[★►\s]+", "", str(value or "")).strip()
    return re.sub(r"\s+", " ", text)


def only_date(value: str) -> str:
    text = str(value or "").strip()
    return text if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text) else ""


def infer_cutoff_type(note: str) -> str:
    if "위탁" in note:
        return "consignment"
    if "송장" in note or "출고완료" in note:
        return "after_shipment"
    return "time"


def infer_cutoff_hour(note: str) -> str:
    match = re.search(r"오(?:전|후)?\s*(\d{1,2})\s*시", str(note or ""))
    if not match:
        match = re.search(r"(\d{1,2})\s*시", str(note or ""))
        if not match:
            return ""
    hour = int(match.group(1))
    note_text = str(note or "")
    if "오후" in note_text and hour < 12:
        hour += 12
    if "오전" in note_text and hour == 12:
        hour = 0
    return f"{hour:02d}" if 0 <= hour <= 23 else ""


def cell_entries(ws, row_start: int = 1, row_end: int = 5, col_end: int = 30):
    entries = []
    for row in range(row_start, row_end + 1):
        for col in range(1, col_end + 1):
            value = ws.cell(row, col).value
            if value in (None, ""):
                continue
            text = str(value).strip()
            if not text:
                continue
            entries.append({"row": row, "col": col, "text": text})
    return entries


def get_text(ws, coord: str) -> str:
    value = ws[coord].value
    return "" if value in (None, "") else str(value).strip()


def find_first(entries, predicate):
    for entry in entries:
        if predicate(entry["text"]):
            return entry["text"]
    return ""


def parse_shipping_rule(text: str):
    note = str(text or "").strip()
    if not note:
        return None

    normalized = note.replace(" ", "")
    fee_regex = r"((?:\d{1,3}(?:,\d{3})+)|(?:\d+\s*(?:천|만))|(?:\d{3,6}))\s*원"

    def parse_fee(raw: str) -> int:
        raw = raw.strip().replace(",", "")
        if raw.endswith("천"):
            return int(raw[:-1]) * 1000
        if raw.endswith("만"):
            return int(raw[:-1]) * 10000
        return int(raw)

    threshold_after_match = re.search(rf"(\d+(?:\.\d+)?)\s*만원\s*(미만|이하)\s*.*?{fee_regex}", note)
    threshold_before_match = re.search(rf"{fee_regex}\s*[/,]?\s*(\d+(?:\.\d+)?)\s*만원\s*(미만|이하)", note)
    free_over_match = re.search(r"(\d+(?:\.\d+)?)\s*만원\s*이상.*?(무료배송|무료|무배)", note)
    amount_match = re.search(r"(\d+(?:\.\d+)?)\s*만원", note)
    fee_tokens = re.findall(fee_regex, note)
    fee_match = fee_tokens[-1] if fee_tokens else ""

    if threshold_after_match:
        threshold_amount = int(float(threshold_after_match.group(1)) * 10000)
        threshold_fee = parse_fee(threshold_after_match.group(3))
        return {
            "shippingPolicyType": "threshold",
            "shippingThresholdAmount": threshold_amount,
            "shippingThresholdFee": threshold_fee,
            "shippingFlatFee": 0,
            "shippingRule": note,
        }

    if threshold_before_match:
        threshold_amount = int(float(threshold_before_match.group(2)) * 10000)
        threshold_fee = parse_fee(threshold_before_match.group(1))
        return {
            "shippingPolicyType": "threshold",
            "shippingThresholdAmount": threshold_amount,
            "shippingThresholdFee": threshold_fee,
            "shippingFlatFee": 0,
            "shippingRule": note,
        }

    if free_over_match and amount_match and fee_match:
        threshold_amount = int(float(amount_match.group(1)) * 10000)
        threshold_fee = parse_fee(fee_match)
        return {
            "shippingPolicyType": "threshold",
            "shippingThresholdAmount": threshold_amount,
            "shippingThresholdFee": threshold_fee,
            "shippingFlatFee": 0,
            "shippingRule": note,
        }

    if "무료배송" in normalized or "무배" in normalized:
        return {
            "shippingPolicyType": "free",
            "shippingThresholdAmount": 0,
            "shippingThresholdFee": 0,
            "shippingFlatFee": 0,
            "shippingRule": note,
        }

    if "무조건" in normalized and fee_match:
        return {
            "shippingPolicyType": "flat",
            "shippingThresholdAmount": 0,
            "shippingThresholdFee": 0,
            "shippingFlatFee": parse_fee(fee_match),
            "shippingRule": note,
        }

    return {
        "shippingPolicyType": "free",
        "shippingThresholdAmount": 0,
        "shippingThresholdFee": 0,
        "shippingFlatFee": 0,
        "shippingRule": note,
    }


def looks_like_shipping_text(text: str) -> bool:
    note = str(text or "").strip()
    if not note:
        return False
    normalized = note.replace(" ", "")
    if any(token in normalized for token in ("배송", "무료", "무배")):
        return True
    if re.search(r"\d+(?:\.\d+)?\s*만원", note) and any(token in normalized for token in ("미만", "이하", "이상")):
        return True
    return False


def infer_shipping_fee_from_rows(ws):
    shipping_col = None
    for col in range(1, 31):
        header = ws.cell(3, col).value
        if str(header or "").strip() == "배송비":
            shipping_col = col
            break
    if not shipping_col:
        return 0

    counter = Counter()
    max_row = ws.max_row or 203
    for row in range(4, min(max_row, 203)):
        value = ws.cell(row, shipping_col).value
        if isinstance(value, (int, float)) and value > 0:
            counter[int(value)] += 1
    return counter.most_common(1)[0][0] if counter else 0


def parse_shipping_from_sheet(ws, row1_2):
    inferred_fee = infer_shipping_fee_from_rows(ws)
    candidates = [get_text(ws, coord) for coord in ("B1", "E1", "G1", "H1")]
    candidates.extend(entry["text"] for entry in row1_2 if looks_like_shipping_text(entry["text"]))
    seen = set()
    for candidate in candidates:
        if not candidate or candidate in seen or not looks_like_shipping_text(candidate):
            continue
        seen.add(candidate)
        shipping = parse_shipping_rule(candidate)
        if shipping and shipping["shippingPolicyType"] == "free" and "만원" in candidate and "이상" in candidate and inferred_fee:
            shipping["shippingPolicyType"] = "threshold"
            amount_match = re.search(r"(\d+(?:\.\d+)?)\s*만원", candidate)
            if amount_match:
                shipping["shippingThresholdAmount"] = int(float(amount_match.group(1)) * 10000)
                shipping["shippingThresholdFee"] = inferred_fee
        if shipping:
            return shipping
    return None


def parse_account_info(text: str):
    note = str(text or "").strip()
    if not note:
        return {}

    bank_name = ""
    for candidate in BANK_NAMES:
        if candidate in note:
            bank_name = candidate
            break

    account_match = re.search(r"(\d[\d-]{5,}\d)", note)
    holder_match = re.search(r"\(([^)]+)\)", note)

    return {
        "bankName": bank_name,
        "bankAccount": account_match.group(1) if account_match else "",
        "accountHolder": holder_match.group(1).strip() if holder_match else "",
    }


def parse_commission_rate(ws, entries):
    for coord in ("L1", "L2", "N2", "E1", "E2"):
        value = ws[coord].value
        if isinstance(value, (int, float)) and 0 < float(value) < 1:
            return round(float(value) * 100, 2)

    percent_pattern = re.compile(r"(\d{1,2}(?:\.\d+)?)\s*%")
    for entry in entries:
        match = percent_pattern.search(entry["text"])
        if match:
            return float(match.group(1))

    margin_pattern = re.compile(r"공급마진\s*\(?\s*(\d{1,2}(?:\.\d+)?)\s*%?\s*\)?")
    for row in range(1, 4):
        for col in range(1, 21):
            value = ws.cell(row, col).value
            if value in (None, ""):
                continue
            text = str(value).strip()
            match = margin_pattern.search(text)
            if match:
                return float(match.group(1))
    return None


def parse_receivable_total(entries):
    labels = ("미지급액 총액", "채권", "채권총액")
    for entry in entries:
        if not any(label in entry["text"] for label in labels):
            continue
        value_entry = next(
            (
                item
                for item in entries
                if item["row"] == entry["row"] and item["col"] == entry["col"] + 1 and re.fullmatch(r"-?\d+(?:\.\d+)?", item["text"])
            ),
            None,
        )
        if value_entry:
            return int(float(value_entry["text"]))
    return None


def infer_settlement_type(entries):
    joined = " ".join(entry["text"] for entry in entries[:20])
    if "위탁" in joined:
        return "consignment"
    if "공급가" in joined or "대리점가" in joined:
        return "prepay_supply"
    return None


def extract_brand_info(ws):
    entries = cell_entries(ws, 1, 5, 30)
    row1_2 = [entry for entry in entries if entry["row"] <= 2]
    result = {}

    cutoff_note = get_text(ws, "C1") or get_text(ws, "D1") or find_first(
        row1_2,
        lambda text: any(token in text for token in ("출고", "송장", "위탁", "월/수/금", "화,목")),
    )
    if cutoff_note:
        result["cutoffNote"] = cutoff_note
        result["cutoffType"] = infer_cutoff_type(cutoff_note)
        result["cutoffHour"] = infer_cutoff_hour(cutoff_note)

    shipping = parse_shipping_from_sheet(ws, row1_2)
    if shipping:
        result.update(shipping)

    business_name = get_text(ws, "R1")
    business_number = get_text(ws, "S1")
    depositor_name = get_text(ws, "T1")
    if business_name:
        result["businessName"] = business_name
    if business_number:
        result["businessNumber"] = business_number
    if depositor_name:
        result["depositorName"] = depositor_name
    if re.fullmatch(r"\d{3}-\d{2}-\d{5}", business_name or "") and business_number and not re.fullmatch(r"\d{3}-\d{2}-\d{5}", business_number or ""):
        result["businessName"] = business_number
        result["businessNumber"] = business_name

    account_holder = get_text(ws, "W1")
    bank_name = get_text(ws, "V1")
    bank_account = get_text(ws, "U1")

    if bank_account or bank_name or account_holder:
        if bank_account and not bank_name:
            parsed = parse_account_info(bank_account)
            bank_account = parsed.get("bankAccount") or bank_account
            bank_name = parsed.get("bankName") or bank_name
            account_holder = parsed.get("accountHolder") or account_holder
        result["bankAccount"] = bank_account
        result["bankName"] = bank_name
        result["accountHolder"] = account_holder or depositor_name
    else:
        account_text = ""
        for coord in ("U1", "K1", "H1", "G1", "E1"):
            if get_text(ws, coord):
                account_text = get_text(ws, coord)
                parsed = parse_account_info(account_text)
                if parsed.get("bankAccount"):
                    result["bankAccount"] = parsed["bankAccount"]
                    result["bankName"] = parsed.get("bankName", "")
                    result["accountHolder"] = parsed.get("accountHolder") or depositor_name
                    break

    commission_rate = parse_commission_rate(ws, row1_2)
    if commission_rate is not None:
        result["commissionRate"] = commission_rate

    settlement_type = infer_settlement_type(row1_2)
    if settlement_type:
        result["settlementType"] = settlement_type

    receivable_total = parse_receivable_total(entries)
    if receivable_total is not None:
        result["receivableTotal"] = receivable_total
        result["hasReceivable"] = receivable_total > 0

    return result


def build_sheet_lookup(workbook):
    lookup = {}
    for name in workbook.sheetnames:
        ws = workbook[name]
        lookup[name] = ws
        lookup[normalize_name(name)] = ws
    return lookup


def main():
    if not DB_PATH.exists():
        raise SystemExit(f"DB file not found: {DB_PATH}")
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook file not found: {WORKBOOK_PATH}")

    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    lookup = build_sheet_lookup(workbook)

    with DB_PATH.open("r", encoding="utf-8") as fh:
      db = json.load(fh)

    backup_path = DATA_DIR / f"db.backup.{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    shutil.copy2(DB_PATH, backup_path)

    updated = []
    for brand in db.get("brands", []):
        if brand.get("type") != "brand":
            continue
        candidates = [
            brand.get("rawSheetName", ""),
            brand.get("name", ""),
            normalize_name(brand.get("rawSheetName", "")),
            normalize_name(brand.get("name", "")),
        ]
        worksheet = next((lookup[key] for key in candidates if key in lookup), None)
        if not worksheet:
            continue
        extracted = extract_brand_info(worksheet)
        if not extracted:
            continue
        changed = {}
        for key, value in extracted.items():
            if value in ("", None):
                continue
            if brand.get(key) != value:
                brand[key] = value
                changed[key] = value
        if changed:
            updated.append((brand.get("name"), changed))

    with DB_PATH.open("w", encoding="utf-8") as fh:
        json.dump(db, fh, ensure_ascii=False, indent=2)

    print(f"Backup: {backup_path.name}")
    print(f"Updated brands: {len(updated)}")
    for name, changed in updated[:40]:
        keys = ", ".join(changed.keys())
        print(f"- {name}: {keys}")


if __name__ == "__main__":
    main()
