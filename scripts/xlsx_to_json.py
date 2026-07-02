#!/usr/bin/env python3
"""Parse an uploaded xlsx to JSON for Node. Emits {"sheets": {name: [rowObjs]}}.

Used for the bank transaction file (은행 거래내역). By default reads the
'통합 라벨링 내역' sheet if present, otherwise the first sheet. Pass --sheet to
override, or --all to dump every sheet. The first non-empty row is the header.
"""

import argparse
import json
import sys
from datetime import date, datetime

from openpyxl import load_workbook


def cell(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    out = []
    for r in rows[1:]:
        if r is None or all(x is None for x in r):
            continue
        obj = {}
        for i, h in enumerate(header):
            obj[h] = cell(r[i]) if i < len(r) else ""
        out.append(obj)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--sheet", default="")
    ap.add_argument("--all", action="store_true")
    args = ap.parse_args()

    wb = load_workbook(args.input, data_only=True, read_only=True)
    result = {}
    if args.all:
        for ws in wb.worksheets:
            result[ws.title] = sheet_rows(ws)
    else:
        target = None
        if args.sheet and args.sheet in wb.sheetnames:
            target = wb[args.sheet]
        elif "통합 라벨링 내역" in wb.sheetnames:
            target = wb["통합 라벨링 내역"]
        else:
            target = wb.worksheets[0]
        result[target.title] = sheet_rows(target)

    json.dump({"sheets": result}, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
