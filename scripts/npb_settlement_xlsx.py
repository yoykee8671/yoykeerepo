#!/usr/bin/env python3
"""Generate the NPB (운영대행) 월별 판매정산서 xlsx, reproducing the
`(도톤)우프_YYYY-M월 판매정산서.xlsx` answer-key layout.

Three sheets:
  1. 종합정산            — rollup row, 재고현황, 정산 방식, 이익분배, 실비 산정표,
                          3PL 단가표, 메모.
  2. 채널별 판매데이터 정리 — one block per channel + grand totals (J78:L82).
  3. DB)입출고목록       — summary rows + the uploaded 입출고 원장 ledger.

CLI: python3 npb_settlement_xlsx.py --input <spec.json> --output <out.xlsx>

Invoked from Node via execFile, mirroring scripts/settlement_excel.py.
See scripts/settlement_excel.py for the shared openpyxl style vocabulary
(Alignment/Border/Font/PatternFill/Side, insert_logo, 맑은 고딕).
"""

import argparse
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    from openpyxl.drawing.image import Image as XLImage
except Exception:  # pragma: no cover
    XLImage = None

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "wooof_logo.png")

# ---------------------------------------------------------------- style vocab
WOOOF_GREEN = "1FA84C"
GREY_FILL = PatternFill("solid", fgColor="D9D9D9")
LIGHT_FILL = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
BLUE_FILL = PatternFill("solid", fgColor="DDEBF7")

_THIN = Side(style="thin", color="808080")
_MED = Side(style="medium", color="404040")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
BORDER_MED = Border(left=_MED, right=_MED, top=_MED, bottom=_MED)

FONT = Font(name="맑은 고딕", size=10)
FONT9 = Font(name="맑은 고딕", size=9)
FONT9B = Font(name="맑은 고딕", size=9, bold=True)
FONT11 = Font(name="맑은 고딕", size=11)
BOLD = Font(name="맑은 고딕", size=10, bold=True)
BOLD11 = Font(name="맑은 고딕", size=11, bold=True)
BOLD12 = Font(name="맑은 고딕", size=12, bold=True)
LOGO_FONT = Font(name="Arial Black", size=20, bold=True, color=WOOOF_GREEN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_NW = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

WON = "#,##0"
PCT = "0%"
DATE_YM = 'yyyy"년"\\ m"월";@'


def insert_logo(ws, cell="A1"):
    """Place the WOOOF logo image; fall back to green text (see settlement_excel.py)."""
    if XLImage and os.path.exists(LOGO_PATH):
        try:
            img = XLImage(LOGO_PATH)
            img.width, img.height = 150, 27
            img.anchor = cell
            ws.add_image(img)
            return
        except Exception:
            pass
    ws[cell] = "WOOOF"
    ws[cell].font = LOGO_FONT
    ws[cell].alignment = LEFT


def box(ws, cell_range, border=BORDER):
    for row in ws[cell_range]:
        for c in row:
            c.border = border


def _num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _fmt_for(v):
    """Infer a display number-format from a value (rates < 1 -> %, big ints -> 천단위)."""
    if isinstance(v, bool):
        return None
    if isinstance(v, float) and not v.is_integer() and 0 < v <= 1:
        return PCT
    if isinstance(v, int) and abs(v) >= 1000:
        return WON
    if isinstance(v, float) and v.is_integer() and abs(v) >= 1000:
        return WON
    return None


# --------------------------------------------------------------- sheet 1
def build_summary(ws, spec):
    ws.sheet_view.showGridLines = False
    period = spec.get("period", {})
    roll = spec.get("rollup", {})

    insert_logo(ws, "A1")

    # --- [판매내역 종합] rollup table (B2:J4) ---
    ws["B2"] = "[판매내역 종합]"
    ws["B2"].font = BOLD
    ws["J2"] = "vat포함"
    ws["J2"].font = FONT11
    ws["J2"].alignment = RIGHT

    heads = ["정산 월", "실판매수량", "판매정가계", "할인계", "실판매계",
             "공제 수수료 \n(공급마진)", "매출계", "실비", "이익"]
    for i, h in enumerate(heads):
        c = ws.cell(row=3, column=2 + i, value=h)
        c.font = BOLD
        c.fill = GREY_FILL
        c.alignment = CENTER
    y = int(period.get("year", 2026))
    m = int(period.get("month", 1))
    dcell = ws.cell(row=4, column=2, value=datetime(y, m, 1))
    dcell.number_format = DATE_YM
    dcell.alignment = CENTER_NW
    dcell.font = FONT
    data = [roll.get("qtyTotal"), roll.get("listTotal"), roll.get("discountTotal"),
            roll.get("realSaleTotal"), roll.get("feeTotal"), roll.get("revenueTotal"),
            roll.get("logisticsCost"), roll.get("profit")]
    for i, v in enumerate(data):
        c = ws.cell(row=4, column=3 + i, value=v)
        c.number_format = WON
        c.alignment = CENTER_NW
        c.font = BOLD11
    box(ws, "B3:J4")

    # --- [재고현황] block (L5:Y12) ---
    ws["L5"] = "[재고현황]"
    ws["L5"].font = BOLD11
    ws["M5"] = "자사물류센터 입출고 기준"
    ws["M5"].font = FONT
    for rng, txt in (("N6:P6", period.get("monthStart", "")),
                     ("Q6:V6", period.get("range", "")),
                     ("W6:Y6", period.get("monthEnd", ""))):
        ws.merge_cells(rng)
        top = rng.split(":")[0]
        ws[top] = txt
        ws[top].font = BOLD11
        ws[top].fill = GREY_FILL
        ws[top].alignment = CENTER_NW
    for rng, txt in (("N7:P7", "기초재고"), ("Q7:S7", "입고"),
                     ("T7:V7", "출고"), ("W7:Y7", "기말재고")):
        ws.merge_cells(rng)
        top = rng.split(":")[0]
        ws[top] = txt
        ws[top].font = BOLD11
        ws[top].fill = LIGHT_FILL
        ws[top].alignment = CENTER_NW
    subheads = ["품목관리코드", "품목명", "전체", "정상", "불용", "전체", "입고",
                "반품", "전체", "판매", "비매출", "전체", "정상", "불용"]
    for i, h in enumerate(subheads):
        c = ws.cell(row=8, column=12 + i, value=h)  # L=12
        c.font = FONT9B
        c.fill = LIGHT_FILL
        c.alignment = CENTER
    inv = spec.get("inventory", [])
    inv_keys = ["openAll", "openOk", "openDead", "inAll", "inIn", "inReturn",
                "outAll", "outSold", "outNonsale", "closeAll", "closeOk", "closeDead"]
    r = 9
    for item in inv:
        ws.cell(row=r, column=12, value=item.get("code")).alignment = CENTER_NW
        ws.cell(row=r, column=13, value=item.get("name")).font = FONT
        for j, k in enumerate(inv_keys):
            c = ws.cell(row=r, column=14 + j, value=item.get(k))
            c.alignment = CENTER_NW
            c.font = FONT
        r += 1
    inv_sum_row = 12
    tot = spec.get("inventoryTotal") or {
        k: sum(_num(it.get(k)) for it in inv) for k in inv_keys
    }
    ws.cell(row=inv_sum_row, column=13, value="합계").font = BOLD
    for j, k in enumerate(inv_keys):
        c = ws.cell(row=inv_sum_row, column=14 + j, value=tot.get(k))
        c.alignment = CENTER_NW
        c.font = BOLD
    box(ws, "L8:Y{}".format(inv_sum_row))

    # --- [정산 방식/ VAT포함] block (B6:J11) ---
    ws["B6"] = "[정산 방식/ VAT포함]"
    ws["B6"].font = BOLD
    ws["B6"].alignment = LEFT
    method_rows = [
        ("항목", "비고"),
        ("A 수수료 (공급마진)", "각 유통 채널별 공급마진 (위탁은 수수료/ 매입은 공급할인액) 기준"),
        ("B 매출계 (월별 최종 판매 금액)",
         "월별 제품 최종판매금액\n(정가 - 프로모션 등 할인이 반영된 최종 결제 금액)"),
        ("C 실비", "운송료 실비, PG결제 수수료 등"),
        ("D 이익",
         "판매 매출 발생 시, “월별 제품 최종 판매 금액”을 기준으로, 인플루언서 공구 "
         "수수료를 포함한 각각의 입점 채널별 공급 수수료, 운송료 실비, PG사 결제 수수료를 제외한 금액"),
    ]
    for i, (label, note) in enumerate(method_rows):
        rr = 7 + i
        ws.merge_cells("B{0}:E{0}".format(rr))
        ws.merge_cells("F{0}:J{0}".format(rr))
        lc = ws["B{}".format(rr)]
        lc.value = label
        lc.font = FONT9B if i == 0 else FONT9
        lc.alignment = CENTER_NW if i == 0 else LEFT
        if i == 1:
            lc.fill = GREEN_FILL
        nc = ws["F{}".format(rr)]
        nc.value = note
        nc.font = FONT9B if i == 0 else FONT9
        nc.alignment = CENTER_NW if i == 0 else LEFT
    box(ws, "B7:J11")

    # --- 이익분배 block (B14:F19) ---
    ws["B14"] = "매출계산서발행"
    ws["B14"].font = BOLD11
    ws["F14"] = "vat포함"
    ws["F14"].font = FONT
    for i, h in enumerate(["항목", "구분", "비율", "금액", "비고"]):
        c = ws.cell(row=15, column=2 + i, value=h)
        c.font = FONT11
        c.fill = LIGHT_FILL
        c.alignment = CENTER_NW
    parties = spec.get("profitSplit", [])
    n = len(parties)
    if n:
        ws.merge_cells("B16:B{}".format(15 + n))
    ws["B16"] = "이익분배"
    ws["B16"].font = FONT11
    ws["B16"].alignment = CENTER
    r = 16
    for p in parties:
        ws.cell(row=r, column=3, value=p.get("partyName")).font = FONT11
        rc = ws.cell(row=r, column=4, value=p.get("ratio"))
        rc.number_format = PCT
        rc.alignment = CENTER_NW
        rc.font = FONT11
        ac = ws.cell(row=r, column=5, value=p.get("amount"))
        ac.number_format = WON
        ac.alignment = CENTER_NW
        ac.font = FONT11
        note = p.get("note") or ("제외" if p.get("excluded") else "")
        ws.cell(row=r, column=6, value=note).font = FONT11
        r += 1
    sum_row = 16 + n
    ws.merge_cells("B{0}:C{0}".format(sum_row))
    ws["B{}".format(sum_row)] = "합계"
    ws["B{}".format(sum_row)].font = FONT11
    ws["B{}".format(sum_row)].alignment = CENTER_NW
    rc = ws.cell(row=sum_row, column=4,
                 value=spec.get("profitSplitTotalRatio", 1))
    rc.number_format = PCT
    rc.alignment = CENTER_NW
    ac = ws.cell(row=sum_row, column=5,
                 value=spec.get("profitSplitTotalAmount", roll.get("profit")))
    ac.number_format = WON
    ac.alignment = CENTER_NW
    box(ws, "B15:F{}".format(sum_row))

    # --- 실비 산정표 (B21:F28) ---
    ws["B21"] = "[실비 : 운임/물류]"
    ws["B21"].font = BOLD
    ws["F21"] = "vat포함"
    ws["F21"].font = FONT
    log = spec.get("logistics", {})
    ws.merge_cells("B22:B23")
    ws["B22"] = "{}년 {}월".format(y, m)
    ws["B22"].font = BOLD
    ws["B22"].alignment = CENTER
    ws.merge_cells("C22:F22")
    ws["C22"] = "물류 실비 산정"
    ws["C22"].font = BOLD
    ws["C22"].alignment = CENTER
    for i, h in enumerate(["수량(출고건수)", "택배", "피킹/패킹/부자재", "총계"]):
        c = ws.cell(row=23, column=3 + i, value=h)
        c.font = BOLD
        c.fill = LIGHT_FILL
        c.alignment = CENTER
    small_count = log.get("smallCount", 0)
    large_count = log.get("largeCount", 0)
    ws["B24"] = "택배(소형)"
    ws["B24"].alignment = CENTER_NW
    ws["C24"] = small_count or None
    ws["D24"] = log.get("smallShip")
    ws["E24"] = log.get("pickPack")
    ws["F24"] = log.get("smallTotal")
    ws["B25"] = "택배(중대형)"
    ws["B25"].alignment = CENTER_NW
    ws["C25"] = large_count or None
    ws["D25"] = log.get("largeShip")
    ws["E25"] = log.get("pickPack")
    ws["F25"] = log.get("largeTotal", 0)
    ws["B26"] = "합계"
    ws["B26"].alignment = CENTER_NW
    ws["B26"].font = BOLD
    ws["C26"] = _num(small_count) + _num(large_count)
    ws["F26"] = log.get("grandTotal", roll.get("logisticsCost"))
    for coord in ("C24", "D24", "E24", "F24", "C25", "D25", "E25", "F25", "C26", "F26"):
        cell = ws[coord]
        cell.number_format = WON
        cell.alignment = CENTER_NW
        if coord in ("C24", "F24", "C26", "F26"):
            cell.font = BOLD11
    box(ws, "B22:F26")
    ws["B27"] = ("*입고비용 및 보관비용은 한시적 청구제외 하였으며, 운송실비(당사가 3PL 지불하는 "
                 "택배운임비 +물류비) 청구기준만 기재되었습니다.")
    ws["B27"].font = FONT9
    ws["B28"] = "*퀵/용달 등 별도 운송비 발생시에는 개별기재합니다."
    ws["B28"].font = FONT9

    # --- 3PL 단가표 (B30:F35) ---
    ws["F29"] = "vat별도"
    ws["F29"].font = FONT
    ws["F29"].alignment = RIGHT
    for i, h in enumerate(["항목", "항목", "견적(원)", "단위", "비고"]):
        c = ws.cell(row=30, column=2 + i, value=h)
        c.font = FONT11
        c.fill = BLUE_FILL
        c.alignment = CENTER_NW
    tpl = spec.get("threePLTable") or [
        {"item": "보관료", "unitPrice": None, "unit": "월/평당 ", "note": "청구제외"},
        {"item": "입고비용", "unitPrice": 0, "unit": "건", "note": "청구제외"},
        {"item": "택배운임비", "unitPrice": 2500, "unit": "건", "note": "로젠택배(소형)"},
        {"item": "택배운임비", "unitPrice": 4000, "unit": "건", "note": "로젠택배(중대형)"},
        {"item": "물류비", "unitPrice": 1300, "unit": "건", "note": "부자재 /피킹/패킹"},
    ]
    ws.merge_cells("B31:B35")
    ws["B31"] = "3PL \n단가표"
    ws["B31"].font = FONT11
    ws["B31"].alignment = CENTER
    for i, row in enumerate(tpl):
        rr = 31 + i
        ws.cell(row=rr, column=3, value=row.get("item")).font = FONT11
        pc = ws.cell(row=rr, column=4, value=row.get("unitPrice"))
        pc.number_format = WON
        pc.alignment = CENTER_NW
        ws.cell(row=rr, column=5, value=row.get("unit")).alignment = CENTER_NW
        ws.cell(row=rr, column=6, value=row.get("note")).font = FONT11
    box(ws, "B30:F35")

    # --- 메모 (B38:B43) ---
    ws["B38"] = "[메모/특이사항]"
    ws["B38"].font = BOLD
    memo = spec.get("memo") or [
        "재고이동(오프라인 위탁진열/ 판매시 정산)",
        "비매출: 협찬, 진열품, 샘플 등의 무상사용",
        "세금계산서는 정산 월 기준 익익월 10일에 발행 예정입니다. ",
        "정산금액은 정산 월 기준 익월 15일에 입금 예정이며 정산금액 활용안 협의에 따라 변경될 수 있습니다. ",
        "불용이슈: 패키지 손상/파손",
    ]
    for i, line in enumerate(memo):
        c = ws.cell(row=39 + i, column=2, value=line)
        c.font = FONT9

    # widths
    for col, w in (("A", 5.2), ("B", 11.3), ("C", 16.7), ("D", 13.3), ("E", 16.3),
                   ("F", 12.0), ("G", 12.3), ("I", 10.5), ("J", 12.3), ("K", 6.8),
                   ("L", 16.2), ("M", 38.3), ("N", 9.5)):
        ws.column_dimensions[col].width = w


# --------------------------------------------------------------- sheet 2
def build_channels(ws, spec):
    ws.sheet_view.showGridLines = False
    period = spec.get("period", {})
    ws["B1"] = "판매기간 {} {} - {}".format(
        period.get("year", ""), period.get("start", ""), period.get("end", ""))
    ws["B1"].font = BOLD11

    r = 4
    for ch in spec.get("channels", []):
        # name row
        ws.cell(row=r, column=2, value=ch.get("name")).font = BOLD11
        if ch.get("desc"):
            ws.cell(row=r, column=4, value=ch["desc"]).font = FONT
        if ch.get("pgTag"):
            ws.cell(row=r, column=7, value=ch["pgTag"]).font = FONT
        if ch.get("tag"):
            tc = ws.cell(row=r, column=11, value=ch["tag"])
            tc.font = FONT
            tc.alignment = CENTER_NW
        hr = r + 1
        headers = ch.get("headers", [])
        for i, h in enumerate(headers):
            c = ws.cell(row=hr, column=2 + i, value=h)
            c.font = BOLD
            c.fill = LIGHT_FILL
            c.alignment = CENTER
        dr = hr + 1
        for row_vals in ch.get("rows", []):
            for i, v in enumerate(row_vals):
                if v is None:
                    continue
                c = ws.cell(row=dr, column=2 + i, value=v)
                c.font = FONT
                c.alignment = CENTER_NW
                fmt = _fmt_for(v)
                if fmt:
                    c.number_format = fmt
            dr += 1
        # 합계 row
        total = ch.get("totalRow")
        if total is not None:
            for i, v in enumerate(total):
                if v is None:
                    continue
                c = ws.cell(row=dr, column=2 + i, value=v)
                c.font = BOLD
                c.alignment = CENTER_NW
                fmt = _fmt_for(v)
                if fmt:
                    c.number_format = fmt
            dr += 1
        box(ws, "B{}:L{}".format(hr, dr - 1))
        r = dr + 1  # one blank row between blocks

    # grand totals (fixed at J78:L82 to match the answer key)
    g = spec.get("grandTotals", {})
    grand = [
        ("총 판매수량", g.get("qty"), "실 수량(1EA단위)"),
        ("정가합계", g.get("list"), "정가 22,000원 기준"),
        ("매출합계", g.get("sales"), "할인/프로모션 반영 판매계"),
        ("공제합계", g.get("deduction"), "공제 수수료 계"),
        ("정산합계", g.get("settle"), "실 정산합계"),
    ]
    for i, (label, val, note) in enumerate(grand):
        rr = 78 + i
        lc = ws.cell(row=rr, column=10, value=label)  # J
        lc.font = BOLD
        lc.fill = LIGHT_FILL
        lc.alignment = CENTER_NW
        vc = ws.cell(row=rr, column=11, value=val)  # K
        vc.number_format = WON
        vc.font = BOLD11
        vc.alignment = CENTER_NW
        nc = ws.cell(row=rr, column=12, value=note)  # L
        nc.font = FONT9
    box(ws, "J78:L82")

    for col, w in (("B", 6), ("C", 30), ("D", 12), ("E", 12), ("F", 11),
                   ("G", 11), ("H", 11), ("I", 12), ("J", 12), ("K", 13), ("L", 20)):
        ws.column_dimensions[col].width = w


# --------------------------------------------------------------- sheet 3
def build_ledger(ws, spec):
    led = spec.get("ledger", {})
    summary = led.get("summary", {})
    # summary rows (C1:I4)
    ws["C1"] = "출고건수 \n(송장기준/건)"
    ws["C1"].font = FONT9B
    ws["C1"].alignment = CENTER
    ws["F1"] = "품목"
    ws["F1"].font = BOLD
    ws["H1"] = "입고합"
    ws["H1"].font = BOLD
    ws["I1"] = "출고합"
    ws["I1"].font = BOLD
    ws["C2"] = summary.get("outCount")
    ws["C2"].font = BOLD11
    ws["C2"].alignment = CENTER_NW
    items = summary.get("items", [])
    r = 2
    for it in items:
        ws.cell(row=r, column=6, value=it.get("name")).font = FONT
        ws.cell(row=r, column=8, value=it.get("inSum")).alignment = CENTER_NW
        ws.cell(row=r, column=9, value=it.get("outSum")).alignment = CENTER_NW
        r += 1
    tr = 2 + len(items)
    ws.cell(row=tr, column=8, value=summary.get("totalIn")).font = BOLD
    ws.cell(row=tr, column=9, value=summary.get("totalOut")).font = BOLD

    # header row A7:T7
    headers = led.get("headers") or [
        "출고카운트", "입출고일자", "입력일자", "공급처명", "바코드번호", "상품명",
        "옵션내용", "원가", "입고수량", "입고금액", "출고수량", "출고금액",
        "재고작업구분", "판매처명", "판매단가", "판매금액", "판매구분", "내용",
        "판매처 주문번호", "판매상세 (판매 수수료율)",
    ]
    for i, h in enumerate(headers):
        c = ws.cell(row=7, column=1 + i, value=h)
        c.font = BOLD
        c.fill = GREY_FILL
        c.alignment = CENTER

    # ledger rows from row 8
    rr = 8
    for row_vals in led.get("rows", []):
        for i, v in enumerate(row_vals):
            if v is None:
                continue
            c = ws.cell(row=rr, column=1 + i, value=v)
            c.font = FONT
            fmt = _fmt_for(v) if not isinstance(v, str) else None
            if fmt:
                c.number_format = fmt
        rr += 1

    widths = [10, 12, 20, 8, 15, 34, 12, 8, 9, 10, 9, 10, 13, 18, 10, 11, 11, 22, 20, 22]
    for i, w in enumerate(widths):
        ws.column_dimensions[ws.cell(row=7, column=1 + i).column_letter].width = w


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    with open(args.input, encoding="utf-8") as f:
        spec = json.load(f)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "종합정산"
    build_summary(ws1, spec)
    build_channels(wb.create_sheet("채널별 판매데이터 정리"), spec)
    build_ledger(wb.create_sheet("DB)입출고목록"), spec)

    wb.save(args.output)
    print(json.dumps({"ok": True,
                      "channels": len(spec.get("channels", [])),
                      "ledgerRows": len(spec.get("ledger", {}).get("rows", []))}))


if __name__ == "__main__":
    main()
