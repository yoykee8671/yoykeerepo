#!/usr/bin/env python3
"""Generate a monthly settlement statement (정산내역서) xlsx matching the
existing wooof template. Reads a JSON spec from --input and writes --output.

Supported settlement types:
  - prepay_debt  (선매입-채권):  commission shown per line but NOT deducted
                                  (미공제) — 납품가합계 = 판매합계.
  - prepay_fee   (선매입-수수료): commission deducted — 납품가합계 = 판매합계 - 수수료.

Spec JSON:
{
  "type": "prepay_debt" | "prepay_fee",
  "supplierName": "KOGONGCAT",
  "year": 2026,
  "monthLabel": "5/1-5/31",
  "rate": 0.25,                     # 계약 수수료율 (fraction)
  "lines": [
    {"itemNo","name","qty","consumer","saleTotal","ship","refundShip",
     "commissionWon","supplyAmt","payDate","note"}, ...
  ],
  "cancels": [
    {"itemNo","name","qty","saleTotal","reason","note"}, ...   # optional
  ]
}
"""

import argparse
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WOOOF_GREEN = "1FA84C"
GREY_FILL = PatternFill("solid", fgColor="D9D9D9")
LIGHT_FILL = PatternFill("solid", fgColor="F2F2F2")

_THIN = Side(style="thin", color="808080")
_MED = Side(style="medium", color="404040")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
BORDER_MED = Border(left=_MED, right=_MED, top=_MED, bottom=_MED)
HEAD_FILL = GREY_FILL

BOLD = Font(bold=True)
FONT = Font(name="맑은 고딕", size=10)
BOLD10 = Font(name="맑은 고딕", size=10, bold=True)
TITLE = Font(name="맑은 고딕", size=26, bold=True)
LOGO_FONT = Font(name="Arial Black", size=30, bold=True, color=WOOOF_GREEN)
NOTICE_BLACK = Font(name="맑은 고딕", size=10, bold=True, color="000000")
NOTICE_RED = Font(name="맑은 고딕", size=10, bold=True, color="C00000")
NOTICE_GREY = Font(name="맑은 고딕", size=10, color="595959")
FINAL_FONT = Font(name="맑은 고딕", size=12, bold=True)
RED_BOLD = Font(name="맑은 고딕", size=10, bold=True, color="C00000")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_NW = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WON = "#,##0"


def box(ws, cell_range, border=BORDER):
    """Apply a border to every cell in an A1:B2 range."""
    for row in ws[cell_range]:
        for c in row:
            c.border = border


def _num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _date(v):
    if not v:
        return ""
    s = str(v)
    # accept ISO / datetime; return YYYY-MM-DD
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[: len(fmt) + 2], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s[:10]


def build_online_debt(ws, lines):
    """선매입-채권 온라인 sheet (13 cols, includes 반품배송비). Returns 합계 row."""
    headers = ["순번", "품목번호", "Product Name", "Qty", "소비자가", "총판매가",
               "배송비", "반품배송비", "수수료(%)", "수수료(원)", "납품가액", "입금일자", "비고"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = BOLD
        c.fill = HEAD_FILL
        c.alignment = CENTER
        c.border = BORDER
    r = 2
    for i, ln in enumerate(lines, start=1):
        vals = [i, ln.get("itemNo", ""), ln.get("name", ""), _num(ln.get("qty")),
                _num(ln.get("consumer")), _num(ln.get("saleTotal")), _num(ln.get("ship")),
                _num(ln.get("refundShip")), None, _num(ln.get("commissionWon")),
                _num(ln.get("supplyAmt")), _date(ln.get("payDate")), ln.get("note", "")]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = BORDER
            if ci in (5, 6, 7, 8, 10, 11):
                c.number_format = WON
        # 수수료(%) as fraction in col I
        ws.cell(row=r, column=9, value=round(_num(ln.get("ratePct", 0)) / 100, 4)).number_format = "0%"
        r += 1
    sr = r  # 합계 row
    ws.cell(row=sr, column=3, value="합계").font = BOLD
    ws.cell(row=sr, column=4, value=f"=SUM(D2:D{sr-1})")
    for col in (6, 7, 8, 10, 11):
        col_letter = ws.cell(row=sr, column=col).column_letter
        cc = ws.cell(row=sr, column=col, value=f"=SUM({col_letter}2:{col_letter}{sr-1})")
        cc.number_format = WON
        cc.font = BOLD
    for ci in range(1, 14):
        ws.cell(row=sr, column=ci).border = BORDER
    widths = [6, 20, 34, 6, 11, 12, 9, 10, 9, 11, 12, 12, 14]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    return sr


def build_online_fee(ws, lines):
    """선매입-수수료 온라인 sheet (12 cols, no 반품배송비). Returns 합계 row."""
    headers = ["순번", "품목번호", "Product Name", "Qty", "소비자가", "총판매가",
               "배송비", "수수료(%)", "수수료(원)", "납품가액", "입금일자", "비고"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = BOLD
        c.fill = HEAD_FILL
        c.alignment = CENTER
        c.border = BORDER
    r = 2
    for i, ln in enumerate(lines, start=1):
        vals = [i, ln.get("itemNo", ""), ln.get("name", ""), _num(ln.get("qty")),
                _num(ln.get("consumer")), _num(ln.get("saleTotal")), _num(ln.get("ship")),
                None, _num(ln.get("commissionWon")), _num(ln.get("supplyAmt")),
                _date(ln.get("payDate")), ln.get("note", "")]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = BORDER
            if ci in (5, 6, 7, 9, 10):
                c.number_format = WON
        ws.cell(row=r, column=8, value=round(_num(ln.get("ratePct", 0)) / 100, 4)).number_format = "0%"
        r += 1
    sr = r
    ws.cell(row=sr, column=3, value="합계").font = BOLD
    ws.cell(row=sr, column=4, value=f"=SUM(D2:D{sr-1})")
    for col in (6, 7, 9, 10):
        col_letter = ws.cell(row=sr, column=col).column_letter
        cc = ws.cell(row=sr, column=col, value=f"=SUM({col_letter}2:{col_letter}{sr-1})")
        cc.number_format = WON
        cc.font = BOLD
    for ci in range(1, 13):
        ws.cell(row=sr, column=ci).border = BORDER
    widths = [6, 20, 34, 6, 11, 12, 9, 9, 11, 12, 12, 14]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    return sr


def build_summary(ws, spec, sr):
    is_debt = spec["type"] == "prepay_debt"
    ws.sheet_view.showGridLines = False

    # --- Header band: WOOOF logo (left) + title (right) ---
    ws.merge_cells("A1:C4")
    ws["A1"] = "WOOOF"
    ws["A1"].font = LOGO_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("F1:H2")
    ws["F1"] = "월별 정산내역서"
    ws["F1"].font = TITLE
    ws["F1"].alignment = Alignment(horizontal="right", vertical="center")

    # --- Notice block (centered, 3 lines) ---
    notices = [
        ("D3:H3", "* 최종 정산금액 기준으로 세금계산서 역발행 예정입니다.", NOTICE_BLACK),
        ("D4:H4", "WEHAGO 사이트를 통해 역발행 예정이오니, 금액 확인 후 '발행 승인' 처리 해주시길 바랍니다.", NOTICE_RED),
        ("D5:H5", "발행 완료 후 업체 '매출'로 국세청 자동 전송됩니다. 감사합니다.", NOTICE_GREY),
    ]
    for rng, text, font in notices:
        ws.merge_cells(rng)
        top = rng.split(":")[0]
        ws[top] = text
        ws[top].font = font
        ws[top].alignment = Alignment(horizontal="center", vertical="center")

    # --- 연도 / 기간 mini table + brand name ---
    ws["B7"] = spec.get("supplierName", "")
    ws["B7"].font = Font(name="맑은 고딕", size=12, bold=True)
    ws["B7"].alignment = LEFT
    for r, label, value in ((6, "연도", spec.get("year")), (7, "기간", spec.get("monthLabel", ""))):
        lc = ws.cell(row=r, column=7, value=label)
        lc.font = BOLD10
        lc.fill = GREY_FILL
        lc.alignment = CENTER_NW
        vc = ws.cell(row=r, column=8, value=value)
        vc.font = FONT
        vc.alignment = CENTER_NW
    box(ws, "G6:H7")

    # --- Main table ---
    comm_header = "판매수수료(미공제)" if is_debt else "판매수수료"
    headers = ["구분", "수수료율", "공급가액", "부가세", "판매합계", comm_header, "납품가합계"]
    for ci, h in enumerate(headers, start=2):  # B..H
        c = ws.cell(row=9, column=ci, value=h)
        c.font = BOLD10
        c.fill = GREY_FILL
        c.alignment = CENTER

    # Row 10 — 온라인 line
    ws["B10"] = "온라인"
    ws["B10"].alignment = CENTER_NW
    ws["B10"].font = FONT
    ws["C10"] = round(_num(spec.get("rate", 0)), 4)
    ws["C10"].number_format = "0%"
    ws["C10"].alignment = CENTER_NW
    ws["D10"] = "=F10/1.1"
    ws["E10"] = "=D10*0.1"
    ws["F10"] = f"=온라인!F{sr}"
    if is_debt:
        ws["G10"] = f"=온라인!J{sr}"   # 수수료(원) col J
        ws["H10"] = f"=온라인!K{sr}"   # 납품가액 col K
    else:
        ws["G10"] = f"=온라인!I{sr}"   # 수수료(원) col I
        ws["H10"] = f"=온라인!J{sr}"   # 납품가액 col J
    for col in ("D", "E", "F", "G", "H"):
        ws[f"{col}10"].number_format = WON
        ws[f"{col}10"].alignment = RIGHT
        ws[f"{col}10"].font = FONT

    # Row 11 — 합계
    ws.merge_cells("B11:E11")
    ws["B11"] = "합계"
    ws["B11"].font = BOLD10
    ws["B11"].alignment = CENTER
    ws["B11"].fill = GREY_FILL
    ws["C11"].fill = GREY_FILL
    ws["D11"].fill = GREY_FILL
    ws["E11"].fill = GREY_FILL
    ws["F11"] = "=SUM(F10)"
    ws["G11"] = 0 if is_debt else "=SUM(G10)"
    ws["H11"] = "=F11-G11"
    for col in ("F", "G", "H"):
        ws[f"{col}11"].number_format = WON
        ws[f"{col}11"].font = BOLD10
        ws[f"{col}11"].alignment = RIGHT
        ws[f"{col}11"].fill = GREY_FILL
    if is_debt:
        ws["G11"].font = RED_BOLD  # 미공제 → 0 in red like the sample
    box(ws, "B9:H11")

    # --- Settlement block ---
    rows = [
        (13, "납품가합계", ("=F11" if is_debt else "=H11")),
        (14, "입금차액 / 기타", None),
        (15, "배송비", f"=온라인!G{sr}"),
        (16, "교환/반품 배송비", (f"=온라인!H{sr}" if is_debt else 0)),
        (17, "최종 정산금액", "=SUM(H13:H16)"),
    ]
    for r, label, value in rows:
        lc = ws.cell(row=r, column=7, value=label)
        lc.fill = GREY_FILL
        lc.alignment = CENTER_NW
        lc.font = BOLD10 if r == 17 else FONT
        vc = ws.cell(row=r, column=8, value=value)
        vc.number_format = WON
        vc.alignment = RIGHT
        vc.font = FINAL_FONT if r == 17 else FONT
    box(ws, "G13:H17")
    ws.cell(row=17, column=7).border = BORDER_MED
    ws.cell(row=17, column=8).border = BORDER_MED

    ws.merge_cells("D19:H19")
    ws["D19"] = "입금차액은 과정산입금 또는 정산입금 후 고객 주문취소 건 등에 해당합니다."
    ws["D19"].font = NOTICE_GREY
    ws["D19"].alignment = Alignment(horizontal="center", vertical="center")

    # --- Column widths & row heights ---
    for col, w in (("A", 12), ("B", 14), ("C", 11), ("D", 11), ("E", 11),
                   ("F", 13), ("G", 18), ("H", 14)):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26
    for r in (9, 10, 11, 13, 14, 15, 16, 17):
        ws.row_dimensions[r].height = 22


def build_cancels(ws, cancels):
    ws["A1"] = "취소/교환 내역 (정산 제외)"
    ws["A1"].font = BOLD
    headers = ["품목번호", "Product Name", "수량", "총판매가", "사유", "비고"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = BOLD
        c.fill = HEAD_FILL
        c.border = BORDER
    r = 3
    for cx in cancels:
        vals = [cx.get("itemNo", ""), cx.get("name", ""), _num(cx.get("qty")),
                _num(cx.get("saleTotal")), cx.get("reason", ""), cx.get("note", "")]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = BORDER
            if ci in (3, 4):
                c.number_format = WON
        r += 1
    for ci, w in enumerate([20, 34, 6, 12, 16, 20], start=1):
        ws.column_dimensions[ws.cell(row=2, column=ci).column_letter].width = w


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    with open(args.input, encoding="utf-8") as f:
        spec = json.load(f)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "총결산"
    ws_online = wb.create_sheet("온라인")

    lines = spec.get("lines", [])
    if spec["type"] == "prepay_debt":
        sr = build_online_debt(ws_online, lines)
    else:
        sr = build_online_fee(ws_online, lines)
    build_summary(ws_sum, spec, sr)

    if spec.get("cancels"):
        build_cancels(wb.create_sheet("취소_교환"), spec["cancels"])

    wb.save(args.output)
    print(json.dumps({"ok": True, "sumRow": sr, "lineCount": len(lines)}))


if __name__ == "__main__":
    main()
